from __future__ import annotations

import math
import uuid
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any, BinaryIO

import pandas as pd
from openpyxl import load_workbook
from supabase import Client, create_client
from supabase.client import ClientOptions


SUPABASE_SCHEMA = "settlement"
IMPORT_TABLE = "excel_import"
PREVIEW_VIEW = "vw_excel_preview"

NORMALIZED_TABLES = (
    "jit_row",
    "penalty_row",
    "atm_balance_row",
    "bonus_route_row",
    "performance_indicator_row",
)


def get_supabase_client(url: str, service_role_key: str) -> Client:
    """Supabase kliens letrehozasa a settlement semahoz."""

    if not url or not url.strip():
        raise ValueError("A Supabase URL nincs megadva.")

    if not service_role_key or not service_role_key.strip():
        raise ValueError("A Supabase SERVICE_ROLE_KEY nincs megadva.")

    return create_client(
        url.strip(),
        service_role_key.strip(),
        options=ClientOptions(
            schema=SUPABASE_SCHEMA,
            postgrest_client_timeout=120,
        ),
    )


def clean_value(value: Any) -> Any:
    """Excel-, pandas- es NumPy-ertek JSON-kompatibilis alakra hozasa."""

    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")

    if hasattr(value, "item"):
        try:
            value = value.item()
        except (ValueError, TypeError):
            pass

    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None

    return value


def _reset_file_pointer(uploaded_file: BinaryIO | BytesIO) -> None:
    try:
        uploaded_file.seek(0)
    except (AttributeError, OSError):
        pass


def _build_raw_row(
    sheet_name: str,
    source_row_no: int,
    values: list[Any],
) -> dict[str, Any] | None:
    cleaned_values = [clean_value(value) for value in values]

    if all(value is None for value in cleaned_values):
        return None

    data = {
        f"column_{column_number}": value
        for column_number, value in enumerate(cleaned_values, start=1)
        if value is not None
    }

    return {
        "sheet_name": sheet_name,
        "source_row_no": source_row_no,
        "data": data,
    }


def _read_xlsx_with_openpyxl(
    uploaded_file: BinaryIO | BytesIO,
) -> list[dict[str, Any]]:
    """XLSX/XLSM osszes sheetjenek nyers beolvasasa."""

    _reset_file_pointer(uploaded_file)

    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=True,
    )

    rows: list[dict[str, Any]] = []

    try:
        for worksheet in workbook.worksheets:
            for source_row_no, row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                raw_row = _build_raw_row(
                    sheet_name=worksheet.title,
                    source_row_no=source_row_no,
                    values=list(row),
                )

                if raw_row is not None:
                    rows.append(raw_row)
    finally:
        workbook.close()

    return rows


def _read_xls_with_pandas(
    uploaded_file: BinaryIO | BytesIO,
) -> list[dict[str, Any]]:
    """
    Regi XLS fajl osszes sheetjenek nyers beolvasasa.

    Ehhez az xlrd csomag szukseges:
        pip install xlrd
    """

    _reset_file_pointer(uploaded_file)

    sheets = pd.read_excel(
        uploaded_file,
        sheet_name=None,
        header=None,
        dtype=object,
        engine="xlrd",
    )

    rows: list[dict[str, Any]] = []

    for sheet_name, dataframe in sheets.items():
        for source_row_no, values in enumerate(
            dataframe.itertuples(index=False, name=None),
            start=1,
        ):
            raw_row = _build_raw_row(
                sheet_name=str(sheet_name),
                source_row_no=source_row_no,
                values=list(values),
            )

            if raw_row is not None:
                rows.append(raw_row)

    return rows


def read_all_excel_sheets(
    uploaded_file: BinaryIO | BytesIO,
) -> list[dict[str, Any]]:
    """Az Excel osszes munkalapjat nyers formaban beolvassa."""

    filename = str(getattr(uploaded_file, "name", "")).lower()

    if filename.endswith(".xls") and not filename.endswith(".xlsx"):
        rows = _read_xls_with_pandas(uploaded_file)
    else:
        rows = _read_xlsx_with_openpyxl(uploaded_file)

    if not rows:
        raise ValueError(
            "Az Excel-fajl egyik munkalapjan sincs nem ures adat."
        )

    return rows


def save_excel_to_supabase(
    uploaded_file: BinaryIO | BytesIO,
    supabase: Client,
    batch_size: int = 500,
) -> dict[str, Any]:
    """Az osszes sheet minden nem ures sorat elmenti a Supabase-be."""

    if batch_size < 1:
        raise ValueError("A batch_size ertekenek legalabb 1-nek kell lennie.")

    raw_rows = read_all_excel_sheets(uploaded_file)
    session_id = str(uuid.uuid4())
    source_file_name = str(getattr(uploaded_file, "name", "") or "")

    import_rows = [
        {
            "session_id": session_id,
            "row_no": import_row_no,
            "sheet_name": row["sheet_name"],
            "source_row_no": row["source_row_no"],
            "data": row["data"],
        }
        for import_row_no, row in enumerate(raw_rows, start=1)
    ]

    inserted_rows = 0

    try:
        for start_index in range(0, len(import_rows), batch_size):
            batch = import_rows[start_index:start_index + batch_size]

            (
                supabase
                .table(IMPORT_TABLE)
                .insert(batch, returning="minimal")
                .execute()
            )

            inserted_rows += len(batch)

    except Exception as exc:
        cleanup_error = None

        try:
            (
                supabase
                .table(IMPORT_TABLE)
                .delete()
                .eq("session_id", session_id)
                .execute()
            )
        except Exception as delete_exc:
            cleanup_error = delete_exc

        message = (
            "Az Excel mentese kozben hiba tortent. "
            "Az ehhez az importhoz korabban elmentett nyers sorokat "
            "megprobaltuk torolni. "
            f"Eredeti hiba: {type(exc).__name__}: {exc}"
        )

        if cleanup_error is not None:
            message += (
                " | Torlesi proba hibaja: "
                f"{type(cleanup_error).__name__}: {cleanup_error}"
            )

        raise RuntimeError(message) from exc

    sheet_names = list(dict.fromkeys(row["sheet_name"] for row in raw_rows))

    sheet_row_counts = {
        sheet_name: sum(
            1 for row in raw_rows if row["sheet_name"] == sheet_name
        )
        for sheet_name in sheet_names
    }

    return {
        "session_id": session_id,
        "source_file_name": source_file_name,
        "inserted_rows": inserted_rows,
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "sheet_row_counts": sheet_row_counts,
    }


def get_import_preview(
    supabase: Client,
    session_id: str,
    sheet_name: str | None = None,
    limit: int = 200,
) -> pd.DataFrame:
    """Egy import legfeljebb limit soranak visszaolvasasa."""

    if limit < 1:
        raise ValueError("A limit ertekenek legalabb 1-nek kell lennie.")

    query = (
        supabase
        .table(PREVIEW_VIEW)
        .select("session_id,row_no,sheet_name,source_row_no,data,created_at")
        .eq("session_id", session_id)
        .order("row_no")
        .range(0, limit - 1)
    )

    if sheet_name:
        query = query.eq("sheet_name", sheet_name)

    response = query.execute()
    records = response.data or []

    preview_rows: list[dict[str, Any]] = []

    for record in records:
        data = record.get("data") or {}

        preview_rows.append(
            {
                "sheet_name": record.get("sheet_name"),
                "source_row_no": record.get("source_row_no"),
                **data,
                "created_at": record.get("created_at"),
            }
        )

    return pd.DataFrame(preview_rows)


def _delete_by_session_id(
    supabase: Client,
    table_name: str,
    session_id: str,
) -> int:
    """Sorok torlese egy tablabol session_id alapjan."""

    response = (
        supabase
        .table(table_name)
        .delete()
        .eq("session_id", session_id)
        .select("id")
        .execute()
    )

    return len(response.data or [])


def delete_excel_import(
    supabase: Client,
    session_id: str,
) -> int:
    """
    Egy teljes import torlese kizarolag a session_id alapjan.

    Onalloan hivhato: nem igenyel Excel-fajlt, nem indit feltoltest,
    es nem indit uj feldolgozast.
    """

    normalized_session_id = str(session_id or "").strip()
    if not normalized_session_id:
        raise ValueError("A session_id megadasa kotelezo a torleshez.")

    deleted_rows = 0

    # 1. Normalizalt adatok
    for table_name in NORMALIZED_TABLES:
        deleted_rows += _delete_by_session_id(
            supabase,
            table_name,
            normalized_session_id,
        )

    # 2. Feldolgozasi gyermekrekordok
    deleted_rows += _delete_by_session_id(
        supabase,
        "validation_error",
        normalized_session_id,
    )
    deleted_rows += _delete_by_session_id(
        supabase,
        "sheet_processing_result",
        normalized_session_id,
    )

    # 3. Feldolgozasi futasok
    deleted_rows += _delete_by_session_id(
        supabase,
        "processing_run",
        normalized_session_id,
    )

    # 4. Nyers import
    deleted_rows += _delete_by_session_id(
        supabase,
        IMPORT_TABLE,
        normalized_session_id,
    )

    return deleted_rows

ZERO_UUID = "00000000-0000-0000-0000-000000000000"


def _delete_all_rows(
    supabase: Client,
    table_name: str,
) -> int:
    """Egy settlement tábla összes, UUID azonosítóval rendelkező sorának törlése."""

    response = (
        supabase
        .table(table_name)
        .delete()
        .neq("id", ZERO_UUID)
        .select("id")
        .execute()
    )

    return len(response.data or [])


def delete_all_settlement_data(supabase: Client) -> dict[str, int]:
    """
    Az importfolyamathoz tartozó settlement táblák teljes kiürítése.

    A művelethez nem kell Excel-fájl vagy session_id, és nem indít
    feltöltést vagy feldolgozást. A törlés gyermek -> szülő sorrendben
    történik, hogy az idegen kulcsok ne akadályozzák a műveletet.
    """

    delete_order = (
        "jit_row",
        "penalty_row",
        "atm_balance_row",
        "bonus_route_row",
        "performance_indicator_row",
        "validation_error",
        "sheet_processing_result",
        "processing_run",
        IMPORT_TABLE,
    )

    deleted_by_table: dict[str, int] = {}

    for table_name in delete_order:
        deleted_by_table[table_name] = _delete_all_rows(
            supabase,
            table_name,
        )

    return deleted_by_table
