import argparse
import os
import re
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import requests


PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from google_client import SCOPES, _load_service_account_info, open_spreadsheet

try:
    import psycopg2
except ImportError:
    psycopg2 = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


SOURCE_NAME = "jitt-workbook"
DEFAULT_SOURCE_SPREADSHEET_ID = "1bplZ8cERp-d0oYjctBkDrZWa-qpxxtpl"
DEFAULT_WORKSHEET_GID = "85404967"
DEFAULT_DETAIL_HEADER_ROW = 23
TABLE_SQL_PATH = PROJECT_ROOT / "docs" / "jitt_workbook_import.sql"


def get_required_env(name):
    value = os.getenv(name, "").strip()

    if not value:
        raise RuntimeError(
            f"Hianyzik a(z) {name} kornyezeti valtozo."
        )

    return value


def get_optional_env(name):
    return os.getenv(name, "").strip()


def build_source_url(source_spreadsheet_id, worksheet_gid=None):
    url = f"https://docs.google.com/spreadsheets/d/{source_spreadsheet_id}/edit"

    if worksheet_gid:
        url = f"{url}?gid={worksheet_gid}#gid={worksheet_gid}"

    return url


def trim_row(row):
    values = ["" if value is None else str(value).strip() for value in row]

    while values and values[-1] == "":
        values.pop()

    return values


def row_has_value(row):
    return any(str(value).strip() for value in row)


def normalize_header(value, index, used):
    text = str(value or "").strip()

    if not text:
        text = f"col_{index}"

    text = text.lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")

    if not text:
        text = f"col_{index}"

    original = text
    suffix = 2

    while text in used:
        text = f"{original}_{suffix}"
        suffix += 1

    used.add(text)
    return text


def normalize_headers(header_row):
    used = set()
    return [
        normalize_header(value, index, used)
        for index, value in enumerate(header_row, start=1)
    ]


def load_native_google_values(source_spreadsheet_id, worksheet_name=None, worksheet_gid=None):
    spreadsheet = open_spreadsheet(source_spreadsheet_id)
    worksheets = spreadsheet.worksheets()

    if worksheet_name:
        worksheet = spreadsheet.worksheet(worksheet_name)
    elif worksheet_gid:
        worksheet = next(
            (
                item
                for item in worksheets
                if str(getattr(item, "id", "")) == str(worksheet_gid)
            ),
            None,
        )

        if worksheet is None:
            raise RuntimeError(
                f"Nincs ilyen gid-u munkalap: {worksheet_gid}"
            )
    else:
        worksheet = worksheets[0]

    return {
        "workbook_title": spreadsheet.title,
        "worksheet_name": worksheet.title,
        "worksheet_gid": str(getattr(worksheet, "id", worksheet_gid or "")),
        "values": [trim_row(row) for row in worksheet.get_all_values()],
    }


def google_authorized_session():
    from google.auth.transport.requests import AuthorizedSession
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(
        _load_service_account_info(),
        scopes=SCOPES,
    )
    return AuthorizedSession(creds)


def download_drive_file_as_xlsx(source_spreadsheet_id):
    session = google_authorized_session()
    metadata_response = session.get(
        f"https://www.googleapis.com/drive/v3/files/{source_spreadsheet_id}",
        params={
            "fields": "id,name,mimeType,size",
            "supportsAllDrives": "true",
        },
        timeout=60,
    )

    if metadata_response.status_code == 404:
        raise RuntimeError(
            "A Drive API nem latja a fajlt. Oszd meg a fajlt a service accounttal, "
            "vagy toltsd le xlsx-kent es hasznald a --xlsx-path opciot."
        )

    metadata_response.raise_for_status()
    metadata = metadata_response.json()
    mime_type = metadata.get("mimeType", "")

    if mime_type == "application/vnd.google-apps.spreadsheet":
        download_response = session.get(
            f"https://www.googleapis.com/drive/v3/files/{source_spreadsheet_id}/export",
            params={
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "supportsAllDrives": "true",
            },
            timeout=120,
        )
    else:
        download_response = session.get(
            f"https://www.googleapis.com/drive/v3/files/{source_spreadsheet_id}",
            params={
                "alt": "media",
                "supportsAllDrives": "true",
            },
            timeout=120,
        )

    download_response.raise_for_status()
    suffix = ".xlsx"
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_file.write(download_response.content)
    temp_file.close()
    return Path(temp_file.name), metadata.get("name") or source_spreadsheet_id


def load_xlsx_values(xlsx_path, worksheet_name=None):
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=xlsx_path,
        data_only=True,
        read_only=True,
    )

    if worksheet_name:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.worksheets[0]

    values = []

    for row in worksheet.iter_rows(values_only=True):
        values.append(trim_row(row))

    return {
        "workbook_title": Path(str(xlsx_path)).stem,
        "worksheet_name": worksheet.title,
        "worksheet_gid": "",
        "values": values,
    }


def load_workbook_values(source_spreadsheet_id, worksheet_name=None, worksheet_gid=None, xlsx_path=None):
    if xlsx_path:
        return load_xlsx_values(
            Path(xlsx_path),
            worksheet_name=worksheet_name,
        )

    try:
        return load_native_google_values(
            source_spreadsheet_id,
            worksheet_name=worksheet_name,
            worksheet_gid=worksheet_gid,
        )
    except Exception as native_error:
        message = str(native_error)

        if "Office file" not in message and "not supported for this document" not in message:
            raise

    downloaded_path, title = download_drive_file_as_xlsx(source_spreadsheet_id)

    try:
        loaded = load_xlsx_values(
            downloaded_path,
            worksheet_name=worksheet_name,
        )
        loaded["workbook_title"] = title
        loaded["worksheet_gid"] = worksheet_gid or ""
        return loaded
    finally:
        try:
            downloaded_path.unlink()
        except OSError:
            pass


def build_rows(loaded, source_spreadsheet_id, source_url, detail_header_row):
    values = loaded["values"]
    imported_at = datetime.now(timezone.utc).isoformat()
    worksheet_name = loaded["worksheet_name"]
    worksheet_gid = loaded.get("worksheet_gid") or ""
    workbook_title = loaded.get("workbook_title") or ""
    header_index = int(detail_header_row) - 1
    header_row = values[header_index] if len(values) > header_index else []
    headers = normalize_headers(header_row)
    main_rows = []
    detail_rows = []

    for index, row in enumerate(values[:header_index], start=1):
        if not row_has_value(row):
            continue

        main_rows.append({
            "source_name": SOURCE_NAME,
            "source_spreadsheet_id": source_spreadsheet_id,
            "source_url": source_url,
            "workbook_title": workbook_title,
            "worksheet_name": worksheet_name,
            "worksheet_gid": worksheet_gid,
            "row_number": index,
            "first_cell": row[0] if row else "",
            "row_values": row,
            "imported_at": imported_at,
            "updated_at": imported_at,
        })

    for index, row in enumerate(values[header_index + 1:], start=detail_header_row + 1):
        if not row_has_value(row):
            continue

        row_data = {
            headers[column_index]: row[column_index] if column_index < len(row) else ""
            for column_index in range(len(headers))
        }

        detail_rows.append({
            "source_name": SOURCE_NAME,
            "source_spreadsheet_id": source_spreadsheet_id,
            "source_url": source_url,
            "workbook_title": workbook_title,
            "worksheet_name": worksheet_name,
            "worksheet_gid": worksheet_gid,
            "row_number": index,
            "row_data": row_data,
            "row_values": row,
            "imported_at": imported_at,
            "updated_at": imported_at,
        })

    import_row = {
        "source_name": SOURCE_NAME,
        "source_spreadsheet_id": source_spreadsheet_id,
        "source_url": source_url,
        "workbook_title": workbook_title,
        "worksheet_name": worksheet_name,
        "worksheet_gid": worksheet_gid,
        "detail_header_row": detail_header_row,
        "detail_headers": headers,
        "top_rows_count": len(main_rows),
        "detail_rows_count": len(detail_rows),
        "imported_at": imported_at,
        "updated_at": imported_at,
    }

    return import_row, main_rows, detail_rows


def ensure_tables_if_possible():
    database_url = get_optional_env("DATABASE_URL")

    if not database_url or psycopg2 is None:
        return

    if not TABLE_SQL_PATH.exists():
        raise RuntimeError(
            f"Hianyzik a tabla SQL fajl: {TABLE_SQL_PATH}"
        )

    with psycopg2.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                TABLE_SQL_PATH.read_text(encoding="utf-8")
            )
            connection.commit()


def raise_for_supabase_error(response):
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        detail = response.text.strip()

        if detail:
            raise requests.HTTPError(
                f"{exc}; Supabase valasz: {detail[:1000]}",
                response=response,
            ) from exc

        raise


def post_supabase_rows(table_name, rows, on_conflict):
    if not rows:
        return 0

    supabase_url = get_required_env("SUPABASE_URL").rstrip("/")
    supabase_key = get_required_env("SUPABASE_SERVICE_ROLE_KEY")
    endpoint = f"{supabase_url}/rest/v1/{table_name}?on_conflict={on_conflict}"
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    response = requests.post(
        endpoint,
        headers=headers,
        json=rows,
        timeout=60,
    )
    raise_for_supabase_error(response)
    return len(rows)


def upsert_rows(import_row, main_rows, detail_rows):
    ensure_tables_if_possible()
    imported = post_supabase_rows(
        "jitt_workbook_imports",
        [import_row],
        "source_name,source_spreadsheet_id,worksheet_name",
    )
    main_count = post_supabase_rows(
        "jitt_workbook_main_raw",
        main_rows,
        "source_name,source_spreadsheet_id,worksheet_name,row_number",
    )
    detail_count = post_supabase_rows(
        "jitt_workbook_detail_raw",
        detail_rows,
        "source_name,source_spreadsheet_id,worksheet_name,row_number",
    )
    return imported, main_count, detail_count


def main():
    parser = argparse.ArgumentParser(
        description="JITT workbook import Google Drive/XLSX forrasbol Supabase DB-be."
    )
    parser.add_argument(
        "--source-id",
        default=DEFAULT_SOURCE_SPREADSHEET_ID,
        help="Google Drive/Sheets file ID.",
    )
    parser.add_argument(
        "--worksheet-gid",
        default=DEFAULT_WORKSHEET_GID,
        help="Google worksheet gid, ha nativ Sheet a forras.",
    )
    parser.add_argument(
        "--worksheet-name",
        default="",
        help="Munkalap neve. XLSX forrasnal ez a biztos valasztas.",
    )
    parser.add_argument(
        "--detail-header-row",
        type=int,
        default=DEFAULT_DETAIL_HEADER_ROW,
        help="Az a sor, amelyik a reszletes tabla fejlece. Alap: 23.",
    )
    parser.add_argument(
        "--xlsx-path",
        default="",
        help="Opcionalis lokalis XLSX fajl. Ha meg van adva, nem hiv Google API-t.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Beolvas es normalizal, de nem ir DB-be.",
    )
    args = parser.parse_args()

    source_url = build_source_url(
        args.source_id,
        args.worksheet_gid,
    )
    loaded = load_workbook_values(
        args.source_id,
        worksheet_name=args.worksheet_name or None,
        worksheet_gid=args.worksheet_gid or None,
        xlsx_path=args.xlsx_path or None,
    )
    import_row, main_rows, detail_rows = build_rows(
        loaded,
        args.source_id,
        source_url,
        args.detail_header_row,
    )

    print(
        f"Forras: {loaded['workbook_title']} / {loaded['worksheet_name']}"
    )
    print(
        f"Felso tabla sorok: {len(main_rows)} | reszletes sorok: {len(detail_rows)}"
    )
    print(
        f"23. sor fejlecek: {', '.join(import_row['detail_headers'][:12])}"
    )

    if detail_rows:
        print(
            f"MINTA row {detail_rows[0]['row_number']}: {detail_rows[0]['row_data']}"
        )

    if args.dry_run:
        print("DRY RUN, DB iras kihagyva.")
        return

    imported, main_count, detail_count = upsert_rows(
        import_row,
        main_rows,
        detail_rows,
    )
    print(
        f"DB feltoltes kesz: imports={imported}, main={main_count}, detail={detail_count}"
    )


if __name__ == "__main__":
    main()
