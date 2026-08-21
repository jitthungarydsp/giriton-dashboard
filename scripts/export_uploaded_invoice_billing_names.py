from __future__ import annotations

import argparse
import base64
import csv
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(PROJECT_ROOT / ".env")
except Exception:
    pass

import pwa_api
from resources.pwa_invoice_validation import parse_invoice_pdf


OUTPUT_DIR = Path("exports")
COMPANY_SUFFIX_RE = re.compile(
    r"\b(kft\.?|bt\.?|zrt\.?|nyrt\.?|ev\.?|e\.v\.?|egy[eé]ni v[aá]llalkoz[oó])\b",
    flags=re.IGNORECASE,
)


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_noise_line(line: str) -> bool:
    text = compact_text(line).casefold()
    if not text:
        return True
    blocked = {
        "szamla",
        "számla",
        "invoice",
        "elado",
        "eladó",
        "szallito",
        "szállító",
        "vevo",
        "vevő",
        "megrendelo",
        "megrendelő",
        "fizetesi mod",
        "fizetési mód",
        "adoszam",
        "adószám",
        "bankszamlaszam",
        "bankszámlaszám",
    }
    if text in blocked:
        return True
    if any(token in text for token in ("adószám", "adoszam", "bankszámla", "bankszamla")):
        return True
    if re.search(r"\d{4}\s+[a-záéíóöőúüű]", text):
        return True
    if re.fullmatch(r"[\d\s./:-]+", text):
        return True
    return False


def candidate_score(line: str) -> int:
    text = compact_text(line)
    score = 0
    if COMPANY_SUFFIX_RE.search(text):
        score += 100
    if re.search(r"[A-ZÁÉÍÓÖŐÚÜŰ][a-záéíóöőúüű]+", text):
        score += 15
    if not re.search(r"\d", text):
        score += 10
    if len(text) <= 80:
        score += 5
    return score


def extract_seller_name_from_text(text: str, seller_tax_number: str = "") -> str:
    lines = [compact_text(line) for line in str(text or "").splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return ""

    label_patterns = ("elado", "eladó", "szallito", "szállító", "kiallito", "kiállító")
    for index, line in enumerate(lines):
        normalized = line.casefold()
        if any(label in normalized for label in label_patterns):
            window = lines[index + 1 : index + 6]
            candidates = [item for item in window if not is_noise_line(item)]
            if candidates:
                return max(candidates, key=candidate_score)

    if seller_tax_number:
        for index, line in enumerate(lines):
            if seller_tax_number in line:
                window = lines[max(0, index - 6) : index]
                candidates = [item for item in window if not is_noise_line(item)]
                if candidates:
                    return max(candidates, key=candidate_score)

    suffix_candidates = [line for line in lines[:30] if not is_noise_line(line) and COMPANY_SUFFIX_RE.search(line)]
    if suffix_candidates:
        return max(suffix_candidates, key=candidate_score)

    candidates = [line for line in lines[:20] if not is_noise_line(line)]
    return max(candidates, key=candidate_score) if candidates else ""


def chunks(items: list[str], size: int = 80) -> list[list[str]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def load_profiles(courier_ids: list[str]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for chunk in chunks([item for item in courier_ids if item]):
        rows = pwa_api.supabase_rest(
            "GET",
            "courier_master",
            params={
                "select": "*",
                "courier_id": f"in.({','.join(chunk)})",
                "limit": str(len(chunk)),
            },
            timeout=60,
        )
        for row in rows or []:
            result[str(row.get("courier_id") or "").strip()] = row
    return result


def load_invoice_documents(limit: int) -> list[dict[str, Any]]:
    return pwa_api.supabase_rest(
        "GET",
        "peopleforce_documents",
        params={
            "select": "id,courier_id,courier_name,document_month,title,file_name,mime_type,file_size,note,uploaded_by,uploaded_at,file_content_base64",
            "document_type": "eq.invoice",
            "order": "uploaded_at.desc",
            "limit": str(limit),
        },
        timeout=120,
    )


def is_cash_invoice_document(row: dict[str, Any]) -> bool:
    title = compact_text(row.get("title")).casefold()
    file_name = compact_text(row.get("file_name")).casefold()
    note = compact_text(row.get("note")).casefold()
    return (
        title.startswith("kp ")
        or "kp számla" in title
        or "kp szamla" in title
        or "kp_szamla" in file_name
        or "fizetési mód: kp" in note
        or "fizetesi mod: kp" in note
    )


def load_tig_documents(courier_ids: list[str], document_months: list[str]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    clean_months = [month for month in document_months if month]
    for courier_chunk in chunks([item for item in courier_ids if item]):
        for month_chunk in chunks(clean_months, size=50):
            rows = pwa_api.supabase_rest(
                "GET",
                "peopleforce_documents",
                params={
                    "select": "id,courier_id,document_month,title,file_name,mime_type,file_size,note,uploaded_at,file_content_base64",
                    "document_type": "eq.tig",
                    "courier_id": f"in.({','.join(courier_chunk)})",
                    "document_month": f"in.({','.join(month_chunk)})",
                    "order": "uploaded_at.desc",
                    "limit": "10000",
                },
                timeout=120,
            )
            result.extend(rows or [])
    return result


def parse_document(row: dict[str, Any]) -> dict[str, Any]:
    parsed: dict[str, Any] = {}
    content_base64 = str(row.get("file_content_base64") or "")
    mime_type = str(row.get("mime_type") or "")
    file_name = str(row.get("file_name") or "")
    if content_base64 and ("pdf" in mime_type.casefold() or file_name.casefold().endswith(".pdf")):
        try:
            content = base64.b64decode(content_base64, validate=True)
            parsed = parse_invoice_pdf(content)
        except Exception as exc:
            parsed = {"parse_error": str(exc)}
    seller_name = extract_seller_name_from_text(
        str(parsed.get("text") or ""),
        str(parsed.get("seller_tax_number") or ""),
    )
    return {
        "invoice_seller_name": seller_name,
        "invoice_seller_tax_number": str(parsed.get("seller_tax_number") or ""),
        "invoice_number": str(parsed.get("invoice_number") or ""),
        "invoice_gross_total_huf": int(parsed.get("gross_total") or 0),
        "parse_error": str(parsed.get("parse_error") or ""),
    }


def parse_huf_amount(value: Any) -> int:
    text = str(value or "")
    text = re.sub(r"([,.])\d{2}\b", "", text)
    digits = re.sub(r"[^0-9-]", "", text)
    try:
        return int(digits)
    except (TypeError, ValueError):
        return 0


def invoice_amount_from_document_note(document: dict[str, Any]) -> int:
    note = str(document.get("note") or "")
    for pattern in [
        r"brutt[óo]\s+[öo]sszesen\s*:?\s*([0-9\s.,]+)\s*Ft",
        r"[öo]sszeg\s*:?\s*([0-9\s.,]+)\s*Ft",
    ]:
        match = re.search(pattern, note, flags=re.IGNORECASE)
        if match:
            return parse_huf_amount(match.group(1))
    return 0


def extract_tig_final_total_from_text(text: str) -> int:
    clean_text = str(text or "")
    patterns = [
        r"v[ée]g[öo]sszeg\s*:?\s*([\d\s.,-]+)\s*ft",
        r"final\s+total\s*:?\s*([\d\s.,-]+)\s*ft",
        r"tig\s+v[ée]g[öo]sszeg\s*:?\s*([\d\s.,-]+)\s*ft",
    ]
    for pattern in patterns:
        match = re.search(pattern, clean_text, flags=re.IGNORECASE)
        if match:
            amount = parse_huf_amount(match.group(1))
            if amount:
                return amount
    amounts = [parse_huf_amount(value) for value in re.findall(r"([\d\s.,-]+)\s*ft", clean_text, flags=re.IGNORECASE)]
    amounts = [amount for amount in amounts if amount]
    return max(amounts) if amounts else 0


def parse_tig_document(row: dict[str, Any]) -> dict[str, Any]:
    content_base64 = str(row.get("file_content_base64") or "")
    mime_type = str(row.get("mime_type") or "")
    file_name = str(row.get("file_name") or "")
    if not content_base64 or ("pdf" not in mime_type.casefold() and not file_name.casefold().endswith(".pdf")):
        return {"tig_final_total_huf": 0, "tig_parse_error": ""}
    try:
        content = base64.b64decode(content_base64, validate=True)
        parsed = parse_invoice_pdf(content)
        return {
            "tig_final_total_huf": extract_tig_final_total_from_text(str(parsed.get("text") or "")),
            "tig_parse_error": "",
        }
    except Exception as exc:
        return {"tig_final_total_huf": 0, "tig_parse_error": str(exc)}


def parse_document_month(value: Any):
    try:
        return pwa_api.parse_month(str(value or ""))
    except Exception:
        return None


def finance_tig_values(user: dict[str, Any], document_month: Any) -> dict[str, Any]:
    month = parse_document_month(document_month)
    if not month:
        return {
            "finance_payable_huf": 0,
            "tig_transfer_gross_huf": 0,
            "tig_final_total_huf": 0,
            "finance_tig_source": "",
            "finance_tig_error": "Hiányzó vagy hibás hónap.",
        }
    try:
        financial = pwa_api.build_financial_breakdown(user, month, allow_unpublished=True)
        tig = pwa_api.build_workflow_tig_breakdown(user, month, financial)
        transfer_gross = 0
        for item in tig.get("rows") or []:
            if str(item.get("key") or "") == "transfer_service":
                transfer_gross = parse_huf_amount(item.get("grossHuf"))
                break
        return {
            "finance_payable_huf": parse_huf_amount(financial.get("totalPayableHuf")),
            "tig_transfer_gross_huf": transfer_gross,
            "tig_final_total_huf": parse_huf_amount(tig.get("finalTotalHuf")),
            "finance_tig_source": str(financial.get("source") or ""),
            "finance_tig_error": "" if financial.get("available") and tig.get("available") else str(financial.get("message") or tig.get("message") or ""),
        }
    except Exception as exc:
        return {
            "finance_payable_huf": 0,
            "tig_transfer_gross_huf": 0,
            "tig_final_total_huf": 0,
            "finance_tig_source": "",
            "finance_tig_error": str(exc),
        }


def latest_invoice_per_courier(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest_rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        courier_id = str(row.get("courier_id") or "").strip()
        if not courier_id or courier_id in seen:
            continue
        seen.add(courier_id)
        latest_rows.append(row)
    return latest_rows


def write_excel(output_path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    def excel_safe(value: Any) -> Any:
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Szamlak"
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    mismatch_fill = PatternFill("solid", fgColor="F4CCCC")
    match_fill = PatternFill("solid", fgColor="D9EAD3")
    for column_index, fieldname in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=column_index, value=fieldname)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_index, row in enumerate(rows, start=2):
        for column_index, fieldname in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=excel_safe(row.get(fieldname, "")))
            if row.get("has_tig_invoice_difference") and fieldname in {
                "payment_tab_amount_huf",
                "payment_tab_tig_final_huf",
                "payment_tab_invoice_amount_huf",
                "payment_tab_difference_huf",
            }:
                cell.fill = mismatch_fill
            elif row.get("has_tig_invoice_match") and fieldname in {
                "payment_tab_amount_huf",
                "payment_tab_tig_final_huf",
                "payment_tab_invoice_amount_huf",
                "payment_tab_difference_huf",
            }:
                cell.fill = match_fill
    for column_index, fieldname in enumerate(fieldnames, start=1):
        values = [str(fieldname)] + [str(row.get(fieldname, "")) for row in rows[:200]]
        width = min(max(len(value) for value in values) + 2, 45)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def write_csv(output_path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def export_uploaded_invoice_billing_names(limit: int, output: Path | None = None, latest_only: bool = False) -> Path:
    rows = [row for row in load_invoice_documents(limit) if not is_cash_invoice_document(row)]
    if latest_only:
        rows = latest_invoice_per_courier(rows)
    courier_ids = sorted({str(row.get("courier_id") or "").strip() for row in rows if row.get("courier_id")})
    profiles = load_profiles(courier_ids)
    suffix = "latest_" if latest_only else ""
    output_path = output or OUTPUT_DIR / f"uploaded_invoice_billing_names_{suffix}{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "courier_id",
        "courier_name_document",
        "profile_courier_name",
        "profile_company_name",
        "invoice_seller_name",
        "profile_tax_number",
        "invoice_seller_tax_number",
        "document_month",
        "invoice_number",
        "payment_tab_amount_huf",
        "payment_tab_tig_final_huf",
        "payment_tab_invoice_amount_huf",
        "payment_tab_difference_huf",
        "file_name",
        "uploaded_at",
        "document_id",
        "needs_company_name_update",
        "has_tig_invoice_match",
        "has_tig_invoice_difference",
        "parse_error",
        "finance_tig_source",
        "finance_tig_error",
    ]
    export_rows = []
    for row in rows:
        courier_id = str(row.get("courier_id") or "").strip()
        profile = profiles.get(courier_id, {})
        parsed = parse_document(row)
        document_month = str(row.get("document_month") or "")[:10]
        user = {
            **profile,
            "courierId": courier_id,
            "username": profile.get("courier_name") or row.get("courier_name"),
        }
        tig_values = finance_tig_values(user, document_month)
        invoice_total = int(parsed.get("invoice_gross_total_huf") or 0) or invoice_amount_from_document_note(row)
        tig_final_total = int(tig_values.get("tig_final_total_huf") or 0)
        finance_payable_total = int(tig_values.get("finance_payable_huf") or 0)
        payment_amount = tig_final_total or finance_payable_total
        difference = invoice_total - payment_amount if invoice_total else 0
        has_match = bool(invoice_total and payment_amount and difference == 0)
        has_difference = not has_match
        profile_company_name = compact_text(profile.get("company_name"))
        invoice_seller_name = compact_text(parsed.get("invoice_seller_name"))
        export_rows.append(
            {
                "courier_id": courier_id,
                "courier_name_document": compact_text(row.get("courier_name")),
                "profile_courier_name": compact_text(profile.get("courier_name")),
                "profile_company_name": profile_company_name,
                "invoice_seller_name": invoice_seller_name,
                "profile_tax_number": compact_text(profile.get("tax_number")),
                "invoice_seller_tax_number": compact_text(parsed.get("invoice_seller_tax_number")),
                "document_month": document_month,
                "invoice_number": compact_text(parsed.get("invoice_number")),
                "payment_tab_amount_huf": payment_amount or "",
                "payment_tab_tig_final_huf": tig_final_total or "",
                "payment_tab_invoice_amount_huf": invoice_total or "",
                "payment_tab_difference_huf": difference if invoice_total else "",
                "file_name": compact_text(row.get("file_name")),
                "uploaded_at": str(row.get("uploaded_at") or ""),
                "document_id": str(row.get("id") or ""),
                "needs_company_name_update": (
                    "yes"
                    if invoice_seller_name and profile_company_name.casefold() != invoice_seller_name.casefold()
                    else ""
                ),
                "has_tig_invoice_match": "yes" if has_match else "",
                "has_tig_invoice_difference": "yes" if has_difference else "",
                "parse_error": compact_text(parsed.get("parse_error")),
                "finance_tig_source": compact_text(tig_values.get("finance_tig_source")),
                "finance_tig_error": compact_text(tig_values.get("finance_tig_error")),
            }
        )
    if output_path.suffix.casefold() == ".csv":
        write_csv(output_path, fieldnames, export_rows)
    else:
        write_excel(output_path, fieldnames, export_rows)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Export uploaded invoice seller names by courier.")
    parser.add_argument("--limit", type=int, default=5000)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--latest-only", action="store_true", help="Only export the newest uploaded invoice for each courier.")
    args = parser.parse_args()
    output_path = export_uploaded_invoice_billing_names(args.limit, args.output, args.latest_only)
    print(f"Export kesz: {output_path}")


if __name__ == "__main__":
    main()
