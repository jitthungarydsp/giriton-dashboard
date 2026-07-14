#!/usr/bin/env python3
"""
Juniusi / havi elso muszak, elso kor keses export mart_dsp_route_stories tablabol.

Pelda:
    python scripts/export_first_shift_delays.py --start-date 2026-06-01 --end-date 2026-06-30
    python scripts/export_first_shift_delays.py --start-date 2026-06-01 --end-date 2026-06-30 --courier-id 7644
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TABLE_NAME = "mart_dsp_route_stories"
RAW_DRIVER_DETAIL_TABLES = [
    "raw_dsp_driver_detail",
    "dsp_driver_detail_raw",
]
PAGE_SIZE = 500

SOURCE_COLUMNS = [
    "work_date",
    "courier_id",
    "courier_name",
    "warehouse_name",
    "route_id",
    "shift_id",
    "shift_name",
    "shift_start",
    "shift_end",
    "available_at",
    "available_for_shift_since",
    "queue_started_at",
    "courier_registered_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
    "queue_entry_delta_minutes",
    "queue_wait_minutes",
    "planned_loading_minutes",
    "real_loading_minutes",
    "planned_route_minutes",
    "real_route_minutes",
    "assigned_to_return_minutes",
    "total_route_minutes",
    "gps_distance_km",
    "checkpoint_straight_km",
    "booking_shift_count",
    "next_booking_shift_text",
    "next_booking_shift_start",
    "next_shift_delay_minutes",
    "address_count",
    "planned_late_count",
    "time_window_late_count",
    "assignment_mode",
]

EXPORT_COLUMNS = [
    "work_date",
    "weekday",
    "courier_id",
    "courier_name",
    "warehouse_name",
    "route_id",
    "shift_id",
    "shift_name",
    "shift_start",
    "shift_end",
    "available_for_shift_since",
    "queue_started_at",
    "courier_registered_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
    "queue_entry_delta_minutes",
    "available_delay_vs_shift_start_minutes",
    "queue_started_delay_vs_shift_start_minutes",
    "queue_wait_minutes",
    "departure_delay_vs_plan_minutes",
    "departure_delay_vs_shift_start_minutes",
    "planned_loading_minutes",
    "real_loading_minutes",
    "planned_route_minutes",
    "real_route_minutes",
    "assigned_to_return_minutes",
    "total_route_minutes",
    "gps_distance_km",
    "address_count",
    "planned_late_count",
    "time_window_late_count",
    "time_window_late_total_minutes",
    "time_window_late_max_minutes",
    "booking_shift_count",
    "next_booking_shift_text",
    "next_shift_delay_minutes",
    "assignment_mode",
    "delay_status",
]

HEADERS_HU = {
    "work_date": "Datum",
    "weekday": "Nap",
    "courier_id": "Futar ID",
    "courier_name": "Futar neve",
    "warehouse_name": "Raktar",
    "route_id": "Elso route ID",
    "shift_id": "Elso muszak ID",
    "shift_name": "Elso muszak",
    "shift_start": "Muszak kezdete",
    "shift_end": "Muszak vege",
    "available_for_shift_since": "Elerheto / sorba allt",
    "queue_started_at": "Sorba allas kezdete",
    "courier_registered_at": "Route regisztracio",
    "assigned_at": "Turat kapott",
    "planned_departure": "Tervezett indulas",
    "real_departure": "Valos indulas",
    "planned_return": "Tervezett vissza",
    "real_return": "Valos vissza",
    "queue_entry_delta_minutes": "Sorbaallas elteres muszakhoz (perc)",
    "available_delay_vs_shift_start_minutes": "Elerheto elteres muszakhoz (perc)",
    "queue_started_delay_vs_shift_start_minutes": "Sorbaallas kezdete elteres muszakhoz (perc)",
    "queue_wait_minutes": "Varakozas turara (perc)",
    "departure_delay_vs_plan_minutes": "Indulas keses tervhez (perc)",
    "departure_delay_vs_shift_start_minutes": "Indulas muszakkezdeshez (perc)",
    "planned_loading_minutes": "Tervezett rakodas (perc)",
    "real_loading_minutes": "Valos rakodas (perc)",
    "planned_route_minutes": "Tervezett turaido (perc)",
    "real_route_minutes": "Valos turaido (perc)",
    "assigned_to_return_minutes": "Kiosztastol visszaig (perc)",
    "total_route_minutes": "Osszes tura hossz (perc)",
    "gps_distance_km": "GPS km",
    "address_count": "Cimek",
    "planned_late_count": "Tervhez keso cimek",
    "time_window_late_count": "Idoablakhoz keso cimek",
    "time_window_late_total_minutes": "Idoablakhoz keses osszesen (perc)",
    "time_window_late_max_minutes": "Legnagyobb idokapu keses (perc)",
    "booking_shift_count": "Foglalas szerinti muszak db",
    "next_booking_shift_text": "Kovetkezo foglalt muszak",
    "next_shift_delay_minutes": "Kov. muszak keses kockazat (perc)",
    "assignment_mode": "Kiosztas modja",
    "delay_status": "Elso kor statusz",
}

SUMMARY_COLUMNS = [
    "courier_id",
    "courier_name",
    "warehouse_names",
    "route_count",
    "address_count",
    "red_queue_delay_count",
    "red_queue_delay_total_minutes",
    "red_queue_delay_max_minutes",
    "time_window_late_count",
    "time_window_late_total_minutes",
    "time_window_late_max_minutes",
    "available_null_count",
]

SUMMARY_HEADERS_HU = {
    "courier_id": "Futar ID",
    "courier_name": "Futar neve",
    "warehouse_names": "Raktarak",
    "route_count": "Ossz kivitt tura",
    "address_count": "Ossz cim",
    "red_queue_delay_count": "Piros sorbaallas/muszak keses db",
    "red_queue_delay_total_minutes": "Piros sorbaallas/muszak keses osszesen (perc)",
    "red_queue_delay_max_minutes": "Legnagyobb piros sorbaallas/muszak keses (perc)",
    "time_window_late_count": "Idoablakhoz keso cimek db",
    "time_window_late_total_minutes": "Idoablakhoz keses osszesen (perc)",
    "time_window_late_max_minutes": "Legnagyobb idokapu keses (perc)",
    "available_null_count": "Elerheto / sorba allt ures db",
}

DATE_COLUMNS = {"work_date"}
DATETIME_COLUMNS = {
    "shift_start",
    "shift_end",
    "available_for_shift_since",
    "queue_started_at",
    "courier_registered_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
}
NUMBER_COLUMNS = {
    "courier_id",
    "route_id",
    "shift_id",
    "queue_entry_delta_minutes",
    "available_delay_vs_shift_start_minutes",
    "queue_started_delay_vs_shift_start_minutes",
    "queue_wait_minutes",
    "departure_delay_vs_plan_minutes",
    "departure_delay_vs_shift_start_minutes",
    "planned_loading_minutes",
    "real_loading_minutes",
    "planned_route_minutes",
    "real_route_minutes",
    "assigned_to_return_minutes",
    "total_route_minutes",
    "gps_distance_km",
    "address_count",
    "planned_late_count",
    "time_window_late_count",
    "time_window_late_total_minutes",
    "time_window_late_max_minutes",
    "booking_shift_count",
    "next_shift_delay_minutes",
}
WEEKDAYS_HU = {
    0: "Hetfo",
    1: "Kedd",
    2: "Szerda",
    3: "Csutortok",
    4: "Pentek",
    5: "Szombat",
    6: "Vasarnap",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Elso muszak / elso kor keses export mart_dsp_route_stories tablabol."
    )
    parser.add_argument("--start-date", required=True, help="Kezdo datum, pl. 2026-06-01")
    parser.add_argument("--end-date", required=True, help="Zaro datum, pl. 2026-06-30")
    parser.add_argument("--courier-id", type=int, help="Csak egy futar exportalasa")
    parser.add_argument("--output", help="Kimeneti Excel fajl")
    parser.add_argument(
        "--page-size",
        type=int,
        default=PAGE_SIZE,
        help=f"Supabase lapmeret, alap: {PAGE_SIZE}",
    )
    return parser.parse_args()


def validate_iso_date(value: str, field_name: str) -> None:
    try:
        date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            f"Hibas {field_name}: {value!r}. Elvart formatum: EEEE-HH-NN."
        ) from exc


def get_supabase_settings() -> tuple[str, str]:
    url = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
    key = (
        os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
        or os.getenv("SUPABASE_KEY", "").strip()
        or os.getenv("SUPABASE_ANON_KEY", "").strip()
    )

    if not url:
        raise RuntimeError("Hianyzik a SUPABASE_URL kornyezeti valtozo.")
    if not key:
        raise RuntimeError(
            "Hianyzik a SUPABASE_SERVICE_ROLE_KEY / SUPABASE_KEY / SUPABASE_ANON_KEY."
        )

    return url, key


def supabase_error_payload(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return response.text


def missing_column_from_response(response: requests.Response) -> str | None:
    if response.status_code != 400:
        return None

    payload = supabase_error_payload(response)
    if not isinstance(payload, dict) or payload.get("code") != "42703":
        return None

    message = str(payload.get("message") or "")
    match = re.search(r"column\s+\S+\.([A-Za-z_][A-Za-z0-9_]*)\s+does not exist", message)
    if match:
        return match.group(1)
    return None


def raise_for_supabase_error(response: requests.Response) -> None:
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        details = supabase_error_payload(response)
        raise requests.HTTPError(
            f"Supabase hiba: HTTP {response.status_code}; url={response.url}; valasz={details}"
        ) from exc


def fetch_raw_driver_detail_rows(
    supabase_url: str,
    supabase_key: str,
    start_date: str,
    end_date: str,
    courier_id: int | None,
    page_size: int,
) -> list[dict[str, Any]]:
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Accept": "application/json",
    }

    last_error: Exception | None = None
    for table_name in RAW_DRIVER_DETAIL_TABLES:
        rows: list[dict[str, Any]] = []
        offset = 0
        endpoint = f"{supabase_url}/rest/v1/{table_name}"

        try:
            while True:
                params: list[tuple[str, str]] = [
                    ("select", "work_date,driver_id,response_json"),
                    ("work_date", f"gte.{start_date}"),
                    ("work_date", f"lte.{end_date}"),
                    ("order", "work_date.asc,driver_id.asc"),
                    ("limit", str(page_size)),
                    ("offset", str(offset)),
                ]
                if courier_id is not None:
                    params.append(("driver_id", f"eq.{courier_id}"))

                response = requests.get(endpoint, headers=headers, params=params, timeout=120)
                raise_for_supabase_error(response)
                page = response.json()

                if not isinstance(page, list):
                    raise RuntimeError(f"Varatlan Supabase valasz: {page!r}")

                rows.extend(page)
                if len(page) < page_size:
                    print(f"Raw idokapu forras: public.{table_name}, sorok: {len(rows)}")
                    return rows
                offset += page_size
        except Exception as exc:  # noqa: BLE001 - next candidate table can still work
            last_error = exc

    print(
        "Figyelmeztetes: nem sikerult raw driver detail sort olvasni, "
        f"az idokapu-keses perc oszlopok uresen maradnak. Utolso hiba: {last_error}",
        file=sys.stderr,
    )
    return []


def fetch_rows(
    supabase_url: str,
    supabase_key: str,
    start_date: str,
    end_date: str,
    courier_id: int | None,
    page_size: int,
) -> list[dict[str, Any]]:
    endpoint = f"{supabase_url}/rest/v1/{TABLE_NAME}"
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Accept": "application/json",
    }
    rows: list[dict[str, Any]] = []
    offset = 0
    select_columns = list(SOURCE_COLUMNS)

    while True:
        params: list[tuple[str, str]] = [
            ("select", ",".join(select_columns)),
            ("work_date", f"gte.{start_date}"),
            ("work_date", f"lte.{end_date}"),
            ("order", "work_date.asc,courier_id.asc,shift_start.asc,assigned_at.asc,route_id.asc"),
            ("limit", str(page_size)),
            ("offset", str(offset)),
        ]
        if courier_id is not None:
            params.append(("courier_id", f"eq.{courier_id}"))

        response = requests.get(endpoint, headers=headers, params=params, timeout=120)
        missing_column = missing_column_from_response(response)
        if missing_column and missing_column in select_columns:
            select_columns.remove(missing_column)
            rows = []
            offset = 0
            print(
                f"Figyelmeztetes: public.{TABLE_NAME}.{missing_column} oszlop nincs az adatbazisban, kihagyom."
            )
            continue

        raise_for_supabase_error(response)
        page = response.json()

        if not isinstance(page, list):
            raise RuntimeError(f"Varatlan Supabase valasz: {page!r}")

        rows.extend(page)
        print(f"Letoltve: {len(page):>4} sor | osszesen: {len(rows):>6} sor")

        if len(page) < page_size:
            break
        offset += page_size

    return rows


def get_route_id(route: dict[str, Any]) -> str:
    return str(route.get("id") or route.get("routeId") or "").strip()


def build_time_window_late_stats(raw_rows: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    stats_by_route: dict[str, dict[str, int]] = {}

    for raw in raw_rows:
        response_json = raw.get("response_json") or {}
        if not isinstance(response_json, dict):
            continue

        for route in response_json.get("routes", []) or []:
            if not isinstance(route, dict):
                continue

            route_id = get_route_id(route)
            if not route_id:
                continue

            route_stats = stats_by_route.setdefault(
                route_id,
                {
                    "time_window_late_total_minutes": 0,
                    "time_window_late_max_minutes": 0,
                },
            )

            for checkpoint in route.get("checkpoints", []) or []:
                if not isinstance(checkpoint, dict):
                    continue

                real_arrival = parse_datetime_value(checkpoint.get("realArrivalTime"))
                window_end = parse_datetime_value(checkpoint.get("deliverTill"))
                if real_arrival is None or window_end is None or real_arrival <= window_end:
                    continue

                late_minutes = int(round((real_arrival - window_end).total_seconds() / 60))
                route_stats["time_window_late_total_minutes"] += late_minutes
                route_stats["time_window_late_max_minutes"] = max(
                    route_stats["time_window_late_max_minutes"],
                    late_minutes,
                )

    return stats_by_route


def parse_datetime_value(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        dt = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    return date.fromisoformat(str(value)[:10])


def minutes_between(later: Any, earlier: Any) -> int | None:
    later_dt = parse_datetime_value(later)
    earlier_dt = parse_datetime_value(earlier)
    if later_dt is None or earlier_dt is None:
        return None
    return int(round((later_dt - earlier_dt).total_seconds() / 60))


def sort_key(row: dict[str, Any]) -> tuple:
    return (
        str(row.get("work_date") or ""),
        int(row.get("courier_id") or 0),
        str(row.get("shift_start") or "9999-99-99T99:99:99"),
        str(row.get("assigned_at") or "9999-99-99T99:99:99"),
        int(row.get("route_id") or 0),
    )


def build_first_shift_rows(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    first_by_day: dict[tuple[str, int], dict[str, Any]] = {}

    for row in sorted(rows, key=sort_key):
        work_date = str(row.get("work_date") or "")[:10]
        courier_id = int(row.get("courier_id") or 0)
        if not work_date or not courier_id:
            continue

        key = (work_date, courier_id)
        if key in first_by_day:
            continue

        first_by_day[key] = dict(row)

    return list(first_by_day.values())


def enrich_route_rows(
    rows: list[dict[str, Any]],
    time_window_stats: dict[str, dict[str, int]] | None = None,
) -> list[dict[str, Any]]:
    time_window_stats = time_window_stats or {}
    enriched_rows: list[dict[str, Any]] = []

    for row in rows:
        output = dict(row)
        route_id = str(row.get("route_id") or "").strip()
        output.update(
            time_window_stats.get(
                route_id,
                {
                    "time_window_late_total_minutes": None,
                    "time_window_late_max_minutes": None,
                },
            )
        )

        work_date = str(row.get("work_date") or "")[:10]
        output["weekday"] = ""
        parsed_date = parse_date_value(work_date)
        if parsed_date is not None:
            output["weekday"] = WEEKDAYS_HU.get(parsed_date.weekday(), "")

        available_delay = minutes_between(
            row.get("available_for_shift_since"),
            row.get("shift_start"),
        )
        queue_started_delay = minutes_between(
            row.get("queue_started_at") or row.get("available_for_shift_since"),
            row.get("shift_start"),
        )
        departure_delay = minutes_between(
            row.get("real_departure"),
            row.get("planned_departure"),
        )
        shift_departure_delay = minutes_between(
            row.get("real_departure"),
            row.get("shift_start"),
        )
        output["available_delay_vs_shift_start_minutes"] = available_delay
        output["queue_started_delay_vs_shift_start_minutes"] = queue_started_delay
        output["departure_delay_vs_plan_minutes"] = departure_delay
        output["departure_delay_vs_shift_start_minutes"] = shift_departure_delay

        queue_delay_for_status = queue_started_delay
        if queue_delay_for_status is None:
            output["delay_status"] = "Nincs elerheto / sorbaallas adat"
        elif queue_delay_for_status > 0:
            output["delay_status"] = f"Sorbaallas keses: {queue_delay_for_status} perc"
        elif queue_delay_for_status < 0:
            output["delay_status"] = f"Idoben/korabban sorba allt: {abs(queue_delay_for_status)} perc"
        else:
            output["delay_status"] = "Pontosan muszakkezdeskor allt sorba"

        enriched_rows.append(output)

    return enriched_rows


def default_output_name(start_date: str, end_date: str, courier_id: int | None) -> Path:
    parts = ["first_shift_delay", start_date, end_date]
    if courier_id is not None:
        parts.append(f"courier_{courier_id}")
    return Path("exports") / ("_".join(parts) + ".xlsx")


def to_number(value: Any, default: float = 0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def build_courier_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[int, dict[str, Any]] = {}

    for row in rows:
        courier_id = int(to_number(row.get("courier_id"), 0))
        if not courier_id:
            continue

        item = summary.setdefault(
            courier_id,
            {
                "courier_id": courier_id,
                "courier_name": row.get("courier_name") or "",
                "warehouse_names_set": set(),
                "route_count": 0,
                "address_count": 0,
                "red_queue_delay_count": 0,
                "red_queue_delay_total_minutes": 0,
                "red_queue_delay_max_minutes": 0,
                "time_window_late_count": 0,
                "time_window_late_total_minutes": 0,
                "time_window_late_max_minutes": 0,
                "available_null_count": 0,
            },
        )

        warehouse_name = str(row.get("warehouse_name") or "").strip()
        if warehouse_name:
            item["warehouse_names_set"].add(warehouse_name)

        item["route_count"] += 1
        item["address_count"] += int(to_number(row.get("address_count"), 0))

        queue_delay = int(
            to_number(
                row.get("queue_started_delay_vs_shift_start_minutes")
                if row.get("queue_started_delay_vs_shift_start_minutes") not in (None, "")
                else row.get("available_delay_vs_shift_start_minutes"),
                0,
            )
        )
        if queue_delay > 0:
            item["red_queue_delay_count"] += 1
            item["red_queue_delay_total_minutes"] += queue_delay
            item["red_queue_delay_max_minutes"] = max(
                item["red_queue_delay_max_minutes"],
                queue_delay,
            )

        item["time_window_late_count"] += int(to_number(row.get("time_window_late_count"), 0))
        item["time_window_late_total_minutes"] += int(
            to_number(row.get("time_window_late_total_minutes"), 0)
        )
        item["time_window_late_max_minutes"] = max(
            item["time_window_late_max_minutes"],
            int(to_number(row.get("time_window_late_max_minutes"), 0)),
        )

        if row.get("available_for_shift_since") in (None, ""):
            item["available_null_count"] += 1

    result = []
    for item in summary.values():
        clean_item = dict(item)
        clean_item["warehouse_names"] = ", ".join(sorted(item["warehouse_names_set"]))
        clean_item.pop("warehouse_names_set", None)
        result.append(clean_item)

    return sorted(result, key=lambda item: (str(item.get("courier_name") or ""), item["courier_id"]))


def write_summary_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
    header_fill: PatternFill,
    header_font: Font,
) -> None:
    worksheet = workbook.create_sheet("Futar statisztika")

    for col_idx, column in enumerate(SUMMARY_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=SUMMARY_HEADERS_HU.get(column, column))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column in enumerate(SUMMARY_COLUMNS, start=1):
            value = row.get(column)
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top")
            if column not in {"courier_name", "warehouse_names"}:
                cell.number_format = "0"

    last_row = max(len(rows) + 1, 2)
    last_col = len(SUMMARY_COLUMNS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    worksheet.row_dimensions[1].height = 42

    widths = {
        "courier_id": 10,
        "courier_name": 24,
        "warehouse_names": 18,
        "route_count": 14,
        "address_count": 12,
        "red_queue_delay_count": 22,
        "red_queue_delay_total_minutes": 28,
        "red_queue_delay_max_minutes": 28,
        "time_window_late_count": 20,
        "time_window_late_total_minutes": 24,
        "time_window_late_max_minutes": 24,
        "available_null_count": 20,
    }
    for col_idx, column in enumerate(SUMMARY_COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(column, 16)


def write_excel(
    first_shift_rows: list[dict[str, Any]],
    all_route_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Elso muszak keses"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    late_fill = PatternFill("solid", fgColor="FFC7CE")
    late_font = Font(color="9C0006")
    early_fill = PatternFill("solid", fgColor="C6EFCE")
    early_font = Font(color="006100")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, column in enumerate(EXPORT_COLUMNS):
        cell = worksheet.cell(row=1, column=col_idx + 1, value=HEADERS_HU.get(column, column))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(first_shift_rows, start=1):
        for col_idx, column in enumerate(EXPORT_COLUMNS):
            value = row.get(column)
            cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
            try:
                if column in DATE_COLUMNS:
                    parsed = parse_date_value(value)
                    cell.value = datetime.combine(parsed, datetime.min.time()) if parsed else None
                    cell.number_format = "yyyy-mm-dd"
                elif column in DATETIME_COLUMNS:
                    parsed_dt = parse_datetime_value(value)
                    cell.value = parsed_dt
                    cell.number_format = "yyyy-mm-dd hh:mm"
                elif column == "gps_distance_km":
                    cell.value = None if value in (None, "") else float(value)
                    cell.number_format = "0.0"
                elif column in NUMBER_COLUMNS:
                    cell.value = None if value in (None, "") else int(float(value))
                    cell.number_format = "0"
                else:
                    cell.value = "" if value is None else str(value)
                cell.alignment = Alignment(vertical="top")
            except (TypeError, ValueError) as exc:
                cell.value = "" if value is None else str(value)
                cell.alignment = Alignment(vertical="top")
                print(
                    f"Figyelmeztetes: sor={row_idx + 1}, oszlop={column}, ertek={value!r}, hiba={exc}",
                    file=sys.stderr,
                )

    last_row = max(len(first_shift_rows) + 1, 2)
    last_col = len(EXPORT_COLUMNS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    worksheet.row_dimensions[1].height = 42

    widths = {
        "work_date": 12,
        "weekday": 12,
        "courier_id": 10,
        "courier_name": 24,
        "warehouse_name": 14,
        "route_id": 14,
        "shift_id": 14,
        "shift_name": 24,
        "shift_start": 18,
        "shift_end": 18,
        "available_for_shift_since": 18,
        "queue_started_at": 18,
        "courier_registered_at": 18,
        "assigned_at": 18,
        "planned_departure": 18,
        "real_departure": 18,
        "planned_return": 18,
        "real_return": 18,
        "queue_entry_delta_minutes": 18,
        "available_delay_vs_shift_start_minutes": 22,
        "queue_started_delay_vs_shift_start_minutes": 24,
        "queue_wait_minutes": 18,
        "departure_delay_vs_plan_minutes": 18,
        "departure_delay_vs_shift_start_minutes": 20,
        "delay_status": 24,
    }
    for col_idx, column in enumerate(EXPORT_COLUMNS):
        worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = widths.get(column, 16)

    for column in [
        "queue_entry_delta_minutes",
        "available_delay_vs_shift_start_minutes",
        "queue_started_delay_vs_shift_start_minutes",
    ]:
        col_letter = get_column_letter(EXPORT_COLUMNS.index(column) + 1)
        range_ref = f"{col_letter}2:{col_letter}{last_row}"
        worksheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator="greaterThan", formula=["0"], fill=late_fill, font=late_font),
        )
        worksheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator="lessThan", formula=["0"], fill=early_fill, font=early_font),
        )

    write_summary_sheet(
        workbook,
        build_courier_summary_rows(all_route_rows),
        header_fill,
        header_font,
    )

    workbook.save(output_path)


def main() -> int:
    args = parse_args()
    validate_iso_date(args.start_date, "--start-date")
    validate_iso_date(args.end_date, "--end-date")
    if args.start_date > args.end_date:
        raise ValueError("A --start-date nem lehet kesobbi, mint a --end-date.")

    supabase_url, supabase_key = get_supabase_settings()
    output_path = (
        Path(args.output)
        if args.output
        else default_output_name(args.start_date, args.end_date, args.courier_id)
    )

    print(f"Forras: public.{TABLE_NAME}")
    print(f"Idoszak: {args.start_date} - {args.end_date}")
    if args.courier_id is not None:
        print(f"Futar ID: {args.courier_id}")

    rows = fetch_rows(
        supabase_url,
        supabase_key,
        args.start_date,
        args.end_date,
        args.courier_id,
        args.page_size,
    )
    raw_rows = fetch_raw_driver_detail_rows(
        supabase_url,
        supabase_key,
        args.start_date,
        args.end_date,
        args.courier_id,
        args.page_size,
    )
    time_window_stats = build_time_window_late_stats(raw_rows)
    enriched_rows = enrich_route_rows(rows, time_window_stats)
    first_shift_rows = build_first_shift_rows(enriched_rows)
    write_excel(first_shift_rows, enriched_rows, output_path)

    print(f"Elso muszak/kor sorok: {len(first_shift_rows)}")
    print(f"Statisztika osszes route sorbol: {len(enriched_rows)}")
    print(f"Kesz: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
