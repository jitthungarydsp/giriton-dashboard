import argparse
import os
import re
import sys
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path

import requests


PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from google_client import SCOPES, _load_service_account_info
from resources.supabase_raw import get_supabase_config, get_supabase_setting

try:
    import psycopg2
except ImportError:
    psycopg2 = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


SOURCE_NAME = "jitt_invoice"
DEFAULT_SOURCE_SPREADSHEET_ID = "1pz9javr4le-J_vmY4oQA4NwdFr8yPx4b"
DEFAULT_DETAIL_HEADER_ROW = 23
TABLE_SQL_PATH = PROJECT_ROOT / "docs" / "jitt_invoice_tables.sql"
BILLING_SHEETS = {"BUD1_JIT", "BUD2_JIT"}


ROUTE_COLUMNS = {
    "location": "location",
    "driver": "driver_name",
    "route_unique_id": "route_unique_id",
    "route_type": "route_type",
    "dsp": "dsp",
    "date": "work_date",
    "orders": "orders",
    "routen": "routes",
    "tip": "tip_huf",
    "license_plate": "license_plate",
    "intern_extern_car": "intern_extern_car",
    "fixed_rate": "fixed_rate_huf",
    "fuel_bonus": "fuel_bonus_huf",
    "car_fridge_bonus": "car_fridge_bonus_huf",
    "branding": "branding_huf",
    "delay_bonus": "delay_bonus_huf",
    "compliance_bonus": "compliance_bonus_huf",
    "fill_rate_bonus": "fill_rate_bonus_huf",
    "comment": "comment",
}

NUMERIC_ROUTE_FIELDS = {
    "orders",
    "routes",
    "tip_huf",
    "fixed_rate_huf",
    "fuel_bonus_huf",
    "car_fridge_bonus_huf",
    "branding_huf",
    "delay_bonus_huf",
    "compliance_bonus_huf",
    "fill_rate_bonus_huf",
}

MONEY_FIELDS = {
    "tip_huf",
    "fixed_rate_huf",
    "fuel_bonus_huf",
    "car_fridge_bonus_huf",
    "branding_huf",
    "delay_bonus_huf",
    "compliance_bonus_huf",
    "fill_rate_bonus_huf",
}


def get_required_env(name):
    value = os.getenv(name, "").strip()

    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")

    return value


def get_optional_env(name):
    value = get_supabase_setting(name)
    return str(value or "").strip()


def build_source_url(source_spreadsheet_id):
    return f"https://docs.google.com/spreadsheets/d/{source_spreadsheet_id}/edit"


def google_authorized_session():
    from google.auth.transport.requests import AuthorizedSession
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(
        _load_service_account_info(),
        scopes=SCOPES,
    )
    return AuthorizedSession(creds)


def download_drive_file_as_xlsx(source_spreadsheet_id):
    session = google_authorized_session()
    metadata_response = session.get(
        f"https://www.googleapis.com/drive/v3/files/{source_spreadsheet_id}",
        params={
            "fields": "id,name,mimeType,size",
            "supportsAllDrives": "true",
        },
        timeout=60,
    )

    if metadata_response.status_code == 404:
        raise RuntimeError(
            "Drive API cannot see the file. Share it with the service account."
        )

    metadata_response.raise_for_status()
    metadata = metadata_response.json()
    mime_type = metadata.get("mimeType", "")

    if mime_type == "application/vnd.google-apps.spreadsheet":
        download_response = session.get(
            f"https://www.googleapis.com/drive/v3/files/{source_spreadsheet_id}/export",
            params={
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "supportsAllDrives": "true",
            },
            timeout=120,
        )
    else:
        download_response = session.get(
            f"https://www.googleapis.com/drive/v3/files/{source_spreadsheet_id}",
            params={
                "alt": "media",
                "supportsAllDrives": "true",
            },
            timeout=120,
        )

    download_response.raise_for_status()
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.write(download_response.content)
    temp_file.close()
    return Path(temp_file.name), metadata.get("name") or source_spreadsheet_id


def clean_text(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def has_value(row):
    return any(clean_text(value) for value in row)


def trim_row(row):
    values = [json_value(value) for value in row]

    while values and values[-1] in ("", None):
        values.pop()

    return values


def json_value(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ")

    if isinstance(value, date):
        return value.isoformat()

    return clean_text(value)


def normalize_header(value, index, used):
    text = clean_text(value).lower()

    if not text:
        text = f"col_{index}"

    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")

    if not text:
        text = f"col_{index}"

    original = text
    suffix = 2

    while text in used:
        text = f"{original}_{suffix}"
        suffix += 1

    used.add(text)
    return text


def normalize_headers(header_row):
    used = set()
    return [
        normalize_header(value, index, used)
        for index, value in enumerate(header_row, start=1)
    ]


def parse_number(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)

    if not text:
        return None

    has_ft = "ft" in text.lower()
    text = (
        text.replace("Ft", "")
        .replace("ft", "")
        .replace("\xa0", "")
        .replace(" ", "")
        .strip()
    )

    if not text:
        return None

    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif has_ft:
        text = text.replace(".", "")

    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)

    if not text:
        return None

    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:19], pattern).date().isoformat()
        except ValueError:
            pass

    return None


def row_to_dict(headers, row):
    return {
        headers[index]: json_value(row[index]) if index < len(row) else ""
        for index in range(len(headers))
    }


def find_route_header_row(worksheet, preferred_row):
    for row_number in range(preferred_row, preferred_row + 5):
        row = next(
            worksheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                values_only=True,
            )
        )
        normalized = normalize_headers(row)

        if "route_unique_id" in normalized and "driver" in normalized:
            return row_number, row, normalized

    raise RuntimeError(
        f"Route header row not found near row {preferred_row} on {worksheet.title}"
    )


def build_summary_rows(worksheet, source_spreadsheet_id, workbook_title, imported_at):
    rows = []

    for row_number in range(5, 23):
        row = next(
            worksheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                values_only=True,
            )
        )

        if not has_value(row):
            continue

        metric_name = clean_text(row[0] if len(row) > 0 else "")

        if not metric_name:
            continue

        rows.append({
            "source_name": SOURCE_NAME,
            "source_spreadsheet_id": source_spreadsheet_id,
            "workbook_title": workbook_title,
            "worksheet_name": worksheet.title,
            "row_number": row_number,
            "metric_name": metric_name,
            "total_value": parse_number(row[1] if len(row) > 1 else None),
            "normal_value": parse_number(row[2] if len(row) > 2 else None),
            "region_value": parse_number(row[3] if len(row) > 3 else None),
            "express_value": parse_number(row[4] if len(row) > 4 else None),
            "total_raw": clean_text(row[1] if len(row) > 1 else ""),
            "normal_raw": clean_text(row[2] if len(row) > 2 else ""),
            "region_raw": clean_text(row[3] if len(row) > 3 else ""),
            "express_raw": clean_text(row[4] if len(row) > 4 else ""),
            "row_values": trim_row(row),
            "imported_at": imported_at,
            "updated_at": imported_at,
        })

    return rows


def build_route_rows(worksheet, source_spreadsheet_id, workbook_title, imported_at, preferred_header_row):
    header_row_number, _header_row, headers = find_route_header_row(
        worksheet,
        preferred_header_row,
    )
    route_rows = []
    final_rows = []

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row_number + 1,
            values_only=True,
        ),
        start=header_row_number + 1,
    ):
        if not has_value(row):
            continue

        row_data = row_to_dict(headers, row)

        if clean_text(row_data.get("route_unique_id")) == "Route Unique ID":
            continue

        record = {
            "source_name": SOURCE_NAME,
            "source_spreadsheet_id": source_spreadsheet_id,
            "workbook_title": workbook_title,
            "worksheet_name": worksheet.title,
            "row_number": row_number,
            "row_data": row_data,
            "row_values": trim_row(row),
            "imported_at": imported_at,
            "updated_at": imported_at,
        }

        for source_key, target_key in ROUTE_COLUMNS.items():
            value = row_data.get(source_key)

            if target_key == "work_date":
                record[target_key] = parse_date(value)
            elif target_key in NUMERIC_ROUTE_FIELDS:
                record[target_key] = parse_number(value)
            else:
                record[target_key] = clean_text(value)

        route_rows.append(record)
        final_rows.append(build_final_route_row(record))

    return route_rows, final_rows


def number_or_zero(value):
    return 0.0 if value is None else float(value)


def build_final_route_row(route_record):
    money_total = sum(
        number_or_zero(route_record.get(field))
        for field in [
            "fuel_bonus_huf",
            "car_fridge_bonus_huf",
            "branding_huf",
            "delay_bonus_huf",
            "compliance_bonus_huf",
            "fill_rate_bonus_huf",
        ]
    )
    without_tip = number_or_zero(route_record.get("fixed_rate_huf")) + money_total
    total = without_tip + number_or_zero(route_record.get("tip_huf"))

    return {
        "source_name": route_record["source_name"],
        "source_spreadsheet_id": route_record["source_spreadsheet_id"],
        "workbook_title": route_record["workbook_title"],
        "worksheet_name": route_record["worksheet_name"],
        "row_number": route_record["row_number"],
        "location": route_record.get("location"),
        "driver_name": route_record.get("driver_name"),
        "route_unique_id": route_record.get("route_unique_id"),
        "route_type": route_record.get("route_type"),
        "dsp": route_record.get("dsp"),
        "work_date": route_record.get("work_date"),
        "orders": route_record.get("orders"),
        "routes": route_record.get("routes"),
        "license_plate": route_record.get("license_plate"),
        "intern_extern_car": route_record.get("intern_extern_car"),
        "fixed_rate_huf": route_record.get("fixed_rate_huf"),
        "fuel_bonus_huf": route_record.get("fuel_bonus_huf"),
        "car_fridge_bonus_huf": route_record.get("car_fridge_bonus_huf"),
        "branding_huf": route_record.get("branding_huf"),
        "delay_bonus_huf": route_record.get("delay_bonus_huf"),
        "compliance_bonus_huf": route_record.get("compliance_bonus_huf"),
        "fill_rate_bonus_huf": route_record.get("fill_rate_bonus_huf"),
        "bonus_total_huf": money_total,
        "tip_huf": route_record.get("tip_huf"),
        "route_total_without_tip_huf": without_tip,
        "route_total_huf": total,
        "comment": route_record.get("comment"),
        "imported_at": route_record["imported_at"],
        "updated_at": route_record["updated_at"],
    }


def build_bonus_route_rows(worksheet, source_spreadsheet_id, workbook_title, imported_at):
    rows = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not has_value(row):
            continue

        rows.append({
            "source_name": SOURCE_NAME,
            "source_spreadsheet_id": source_spreadsheet_id,
            "workbook_title": workbook_title,
            "worksheet_name": worksheet.title,
            "row_number": row_number,
            "dsp": clean_text(row[0] if len(row) > 0 else ""),
            "site": clean_text(row[1] if len(row) > 1 else ""),
            "courier_id": clean_text(row[2] if len(row) > 2 else ""),
            "driver_name": clean_text(row[3] if len(row) > 3 else ""),
            "routes": parse_number(row[4] if len(row) > 4 else None),
            "bonus_huf": parse_number(row[5] if len(row) > 5 else None),
            "row_values": trim_row(row),
            "imported_at": imported_at,
            "updated_at": imported_at,
        })

    return rows


def build_penalty_rows(worksheet, source_spreadsheet_id, workbook_title, imported_at):
    rows = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not has_value(row):
            continue

        rows.append({
            "source_name": SOURCE_NAME,
            "source_spreadsheet_id": source_spreadsheet_id,
            "workbook_title": workbook_title,
            "worksheet_name": worksheet.title,
            "row_number": row_number,
            "penalty_type": clean_text(row[0] if len(row) > 0 else ""),
            "penalty_date": parse_date(row[1] if len(row) > 1 else None),
            "courier_id": clean_text(row[2] if len(row) > 2 else ""),
            "driver_name": clean_text(row[3] if len(row) > 3 else ""),
            "dsp": clean_text(row[4] if len(row) > 4 else ""),
            "site": clean_text(row[5] if len(row) > 5 else ""),
            "note": clean_text(row[6] if len(row) > 6 else ""),
            "amount_huf": parse_number(row[7] if len(row) > 7 else None),
            "extra_note": clean_text(row[8] if len(row) > 8 else ""),
            "row_values": trim_row(row),
            "imported_at": imported_at,
            "updated_at": imported_at,
        })

    return rows


def build_contract_rule_rows(imported_at):
    durations = [
        (2.0, "<= 2.0 ora / Expressz", [1333, 666, 333]),
        (3.0, "3.0 ora", [2000, 1000, 500]),
        (3.5, "3.5 ora", [2333, 1166, 583]),
        (4.0, "4.0 ora", [2666, 1333, 666]),
        (4.5, "4.5 ora / Standard varosi alap", [3000, 1500, 750]),
        (5.0, "5.0 ora", [3333, 1666, 833]),
        (5.5, "5.5 ora / Regionalis alap", [3666, 1833, 916]),
        (6.0, "6.0 ora", [4000, 2000, 1000]),
    ]
    metrics = [
        (
            "delay",
            "Kesesdelmi mutato",
            [
                (1, "Szint 1", "<= 1.50%", None, 1.50),
                (2, "Szint 2", "1.51% - 3.00%", 1.51, 3.00),
                (3, "Szint 3", "3.01% - 5.00%", 3.01, 5.00),
            ],
        ),
        (
            "tour_compliance",
            "Turamegfelelesi mutato",
            [
                (1, "Szint 1", "<= 2.00%", None, 2.00),
                (2, "Szint 2", "2.01% - 4.00%", 2.01, 4.00),
                (3, "Szint 3", "4.01% - 10.00%", 4.01, 10.00),
            ],
        ),
    ]
    rows = []

    for metric_type, metric_name, levels in metrics:
        for level_number, level_name, threshold_label, min_pct, max_pct in levels:
            for duration_hours, duration_label, amounts in durations:
                amount_huf = amounts[level_number - 1]
                rule_id = f"{metric_type}_{level_number}_{str(duration_hours).replace('.', '_')}"
                rows.append({
                    "rule_id": rule_id,
                    "metric_type": metric_type,
                    "metric_name": metric_name,
                    "level_number": level_number,
                    "level_name": level_name,
                    "threshold_label": threshold_label,
                    "threshold_min_pct": min_pct,
                    "threshold_max_pct": max_pct,
                    "duration_hours": duration_hours,
                    "duration_label": duration_label,
                    "amount_huf": amount_huf,
                    "source_note": "Contract screenshots provided by user, net HUF amounts.",
                    "imported_at": imported_at,
                    "updated_at": imported_at,
                })

    return rows


def build_rows(source_spreadsheet_id, xlsx_path, detail_header_row):
    from openpyxl import load_workbook

    imported_at = datetime.now(timezone.utc).isoformat()
    workbook_path = Path(xlsx_path)
    workbook = load_workbook(
        filename=workbook_path,
        read_only=True,
        data_only=True,
    )
    workbook_title = workbook_path.stem
    imports = [{
        "source_name": SOURCE_NAME,
        "source_spreadsheet_id": source_spreadsheet_id,
        "source_url": build_source_url(source_spreadsheet_id),
        "workbook_title": workbook_title,
        "imported_at": imported_at,
        "updated_at": imported_at,
    }]
    summary_rows = []
    route_rows = []
    final_rows = []
    bonus_rows = []
    penalty_rows = []

    try:
        for worksheet in workbook.worksheets:
            if worksheet.title in BILLING_SHEETS:
                summary_rows.extend(
                    build_summary_rows(
                        worksheet,
                        source_spreadsheet_id,
                        workbook_title,
                        imported_at,
                    )
                )
                worksheet_route_rows, worksheet_final_rows = build_route_rows(
                    worksheet,
                    source_spreadsheet_id,
                    workbook_title,
                    imported_at,
                    detail_header_row,
                )
                route_rows.extend(worksheet_route_rows)
                final_rows.extend(worksheet_final_rows)
            elif worksheet.title == "Bonus routes":
                bonus_rows.extend(
                    build_bonus_route_rows(
                        worksheet,
                        source_spreadsheet_id,
                        workbook_title,
                        imported_at,
                    )
                )
            elif worksheet.title == "Penalties":
                penalty_rows.extend(
                    build_penalty_rows(
                        worksheet,
                        source_spreadsheet_id,
                        workbook_title,
                        imported_at,
                    )
                )
    finally:
        workbook.close()

    contract_rows = build_contract_rule_rows(imported_at)

    return {
        "imports": imports,
        "summary": summary_rows,
        "routes": route_rows,
        "final": final_rows,
        "bonus": bonus_rows,
        "penalties": penalty_rows,
        "contract": contract_rows,
    }


def ensure_tables_if_possible():
    database_url = get_optional_env("DATABASE_URL")

    if not database_url or psycopg2 is None:
        return False

    with psycopg2.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(TABLE_SQL_PATH.read_text(encoding="utf-8"))
            connection.commit()

    return True


def raise_for_supabase_error(response):
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        detail = response.text.strip()

        if detail:
            raise requests.HTTPError(
                f"{exc}; Supabase response: {detail[:1000]}",
                response=response,
            ) from exc

        raise


def post_supabase_rows(table_names, rows, on_conflict):
    if not rows:
        return 0

    if isinstance(table_names, str):
        table_names = [table_names]

    supabase_url, supabase_key = get_supabase_config()

    if not supabase_url or not supabase_key:
        raise RuntimeError(
            "Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY setting."
        )

    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    total = 0
    selected_table = None

    for index in range(0, len(rows), 500):
        batch = rows[index:index + 500]
        response = None

        for table_name in ([selected_table] if selected_table else table_names):
            if not table_name:
                continue

            endpoint = f"{supabase_url}/rest/v1/{table_name}?on_conflict={on_conflict}"
            response = requests.post(
                endpoint,
                headers=headers,
                json=batch,
                timeout=90,
            )

            if response.status_code in [404, 406]:
                continue

            raise_for_supabase_error(response)
            selected_table = table_name
            break
        else:
            if response is not None:
                raise_for_supabase_error(response)

            raise RuntimeError(
                f"No writable invoice table found: {', '.join(table_names)}"
            )

        total += len(batch)

    return total


def upsert_all(rows):
    ensure_tables_if_possible()
    return {
        "imports": post_supabase_rows(
            ["bill_jitt_invoice_imports", "jitt_invoice_imports"],
            rows["imports"],
            "source_name,source_spreadsheet_id",
        ),
        "summary": post_supabase_rows(
            ["bill_jitt_invoice_summary", "jitt_invoice_summary_rows"],
            rows["summary"],
            "source_name,source_spreadsheet_id,worksheet_name,row_number",
        ),
        "routes": post_supabase_rows(
            ["bill_jitt_invoice_routes", "jitt_invoice_route_rows"],
            rows["routes"],
            "source_name,source_spreadsheet_id,worksheet_name,row_number",
        ),
        "final": post_supabase_rows(
            ["bill_jitt_invoice_final_routes", "jitt_invoice_final_routes"],
            rows["final"],
            "source_name,source_spreadsheet_id,worksheet_name,row_number",
        ),
        "bonus": post_supabase_rows(
            ["bill_jitt_invoice_bonus_routes", "jitt_invoice_bonus_routes"],
            rows["bonus"],
            "source_name,source_spreadsheet_id,worksheet_name,row_number",
        ),
        "penalties": post_supabase_rows(
            ["bill_jitt_invoice_penalties", "jitt_invoice_penalties"],
            rows["penalties"],
            "source_name,source_spreadsheet_id,worksheet_name,row_number",
        ),
        "contract": post_supabase_rows(
            ["bill_jitt_contract_bonus_rules", "jitt_invoice_contract_bonus_rules"],
            rows["contract"],
            "rule_id",
        ),
    }


def print_counts(rows):
    print(
        "ROWS "
        f"imports={len(rows['imports'])} "
        f"summary={len(rows['summary'])} "
        f"routes={len(rows['routes'])} "
        f"final={len(rows['final'])} "
        f"bonus={len(rows['bonus'])} "
        f"penalties={len(rows['penalties'])} "
        f"contract={len(rows['contract'])}"
    )

    if rows["routes"]:
        sample = rows["routes"][0]
        print(
            "SAMPLE_ROUTE "
            f"{sample.get('worksheet_name')} "
            f"{sample.get('driver_name')} "
            f"{sample.get('route_unique_id')} "
            f"{sample.get('work_date')}"
        )


def main():
    parser = argparse.ArgumentParser(
        description="Load JITT invoice workbook into jitt_invoice_* Supabase tables."
    )
    parser.add_argument(
        "--source-id",
        default=DEFAULT_SOURCE_SPREADSHEET_ID,
        help="Google Drive file ID.",
    )
    parser.add_argument(
        "--xlsx-path",
        default="",
        help="Optional local xlsx path. If omitted, the script downloads from Drive.",
    )
    parser.add_argument(
        "--detail-header-row",
        type=int,
        default=DEFAULT_DETAIL_HEADER_ROW,
        help="Preferred route detail header row. Default: 23.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse only; do not write to DB.",
    )
    args = parser.parse_args()

    downloaded_path = None

    if args.xlsx_path:
        workbook_path = Path(args.xlsx_path)
    else:
        downloaded_path, title = download_drive_file_as_xlsx(args.source_id)
        workbook_path = downloaded_path
        print(f"Downloaded workbook: {title}")

    try:
        rows = build_rows(
            args.source_id,
            workbook_path,
            args.detail_header_row,
        )
        print_counts(rows)

        if args.dry_run:
            print("DRY_RUN no DB write.")
            return

        result = upsert_all(rows)
        print(f"DB_UPSERT {result}")
    finally:
        if downloaded_path:
            try:
                downloaded_path.unlink()
            except OSError:
                pass


if __name__ == "__main__":
    main()
