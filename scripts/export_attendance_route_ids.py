#!/usr/bin/env python3
"""
fetch-attendance route ID export Excelbe.

Pelda:
    python scripts/export_attendance_route_ids.py --month 2026-07
    python scripts/export_attendance_route_ids.py --start-date 2026-07-01 --end-date 2026-07-31
"""

from __future__ import annotations

import argparse
import calendar
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from scripts.dsp_incremental_common import (
        get_supabase_config,
        is_missing_table_response,
        raise_for_supabase_error,
        supabase_headers,
    )
except ModuleNotFoundError:
    from dsp_incremental_common import (
        get_supabase_config,
        is_missing_table_response,
        raise_for_supabase_error,
        supabase_headers,
    )


ROUTE_TABLE_CANDIDATES = [
    "dsp_attendance_routes",
    "stg_dsp_attendance_routes",
]
RAW_TABLE_CANDIDATES = [
    "raw_dsp_attendance",
    "dsp_attendance_raw",
]
PAGE_SIZE = 500

ROUTE_COLUMNS = [
    "work_date",
    "courier_id",
    "courier_name",
    "warehouse_name",
    "route_id",
    "courier_registered_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
    "planned_route_minutes",
    "real_route_minutes",
    "departure_diff_minutes",
    "return_diff_minutes",
    "departure_status",
    "return_status",
]

EXPORT_COLUMNS = [
    "work_date",
    "courier_id",
    "courier_name",
    "warehouse_name",
    "route_id",
    "courier_registered_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
    "planned_route_minutes",
    "real_route_minutes",
    "departure_diff_minutes",
    "return_diff_minutes",
    "departure_status",
    "return_status",
    "source_table",
]

HEADERS_HU = {
    "work_date": "Datum",
    "courier_id": "Futar ID",
    "courier_name": "Futar neve",
    "warehouse_name": "Raktar",
    "route_id": "Route ID",
    "courier_registered_at": "Futar regisztralt",
    "assigned_at": "Turat kapott",
    "planned_departure": "Tervezett indulas",
    "real_departure": "Valos indulas",
    "planned_return": "Tervezett visszaerkezes",
    "real_return": "Valos visszaerkezes",
    "planned_route_minutes": "Tervezett turahossz (perc)",
    "real_route_minutes": "Valos turahossz (perc)",
    "departure_diff_minutes": "Indulas elteres (perc)",
    "return_diff_minutes": "Visszaerkezes elteres (perc)",
    "departure_status": "Indulas statusz",
    "return_status": "Visszaerkezes statusz",
    "source_table": "Forras tabla",
}

DATETIME_COLUMNS = {
    "courier_registered_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
}
DATE_COLUMNS = {"work_date"}
NUMBER_COLUMNS = {
    "courier_id",
    "route_id",
    "planned_route_minutes",
    "real_route_minutes",
    "departure_diff_minutes",
    "return_diff_minutes",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Havi fetch-attendance route ID export Excelbe."
    )
    parser.add_argument("--month", help="Honap YYYY-MM formatumban, pl. 2026-07")
    parser.add_argument("--start-date", help="Kezdo datum YYYY-MM-DD")
    parser.add_argument("--end-date", help="Zaro datum YYYY-MM-DD")
    parser.add_argument("--output", help="Kimeneti Excel fajl")
    parser.add_argument("--page-size", type=int, default=PAGE_SIZE)
    parser.add_argument(
        "--source",
        choices=["auto", "routes", "raw"],
        default="auto",
        help="Forras: auto = teljes honapot lehetoleg bontott tablabol, hianyos bontott tabla eseten raw JSON.",
    )
    return parser.parse_args()


def month_bounds(month_text: str) -> tuple[str, str]:
    year, month = map(int, month_text.split("-"))
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


def resolve_dates(args: argparse.Namespace) -> tuple[str, str, str]:
    if args.month:
        start_date, end_date = month_bounds(args.month)
        return args.month, start_date, end_date

    if not args.start_date:
        raise ValueError("Adj meg --month vagy --start-date erteket.")

    start_date = args.start_date
    end_date = args.end_date or args.start_date
    period_label = start_date[:7] if start_date[:7] == end_date[:7] else f"{start_date}_{end_date}"
    return period_label, start_date, end_date


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def parse_datetime_value(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        except ValueError:
            return None
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt


def to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def minutes_between(later: Any, earlier: Any) -> int | None:
    later_dt = parse_datetime_value(later)
    earlier_dt = parse_datetime_value(earlier)
    if later_dt is None or earlier_dt is None:
        return None
    return int(round((later_dt - earlier_dt).total_seconds() / 60))


def request_table_page(
    table_name: str,
    params: list[tuple[str, str]],
) -> requests.Response:
    supabase_url, _service_role_key = get_supabase_config()
    endpoint = f"{supabase_url}/rest/v1/{table_name}"
    return requests.get(
        endpoint,
        headers=supabase_headers({"Accept": "application/json"}),
        params=params,
        timeout=120,
    )


def fetch_table_rows(
    table_name: str,
    *,
    select: str,
    start_date: str,
    end_date: str,
    page_size: int,
    order: str,
) -> list[dict[str, Any]] | None:
    rows: list[dict[str, Any]] = []
    offset = 0

    while True:
        params = [
            ("select", select),
            ("work_date", f"gte.{start_date}"),
            ("work_date", f"lte.{end_date}"),
            ("order", order),
            ("limit", str(page_size)),
            ("offset", str(offset)),
        ]
        response = request_table_page(table_name, params)
        if is_missing_table_response(response):
            return None

        raise_for_supabase_error(response, table_name)
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError(f"Varatlan Supabase valasz: {page!r}")

        rows.extend(page)
        print(f"{table_name}: {len(page):>4} sor | osszesen: {len(rows):>6}")
        if len(page) < page_size:
            break
        offset += page_size

    return rows


def read_route_table_rows(
    *,
    start_date: str,
    end_date: str,
    page_size: int,
) -> tuple[str, list[dict[str, Any]]]:
    for table_name in ROUTE_TABLE_CANDIDATES:
        rows = fetch_table_rows(
            table_name,
            select=",".join(ROUTE_COLUMNS),
            start_date=start_date,
            end_date=end_date,
            page_size=page_size,
            order="work_date.asc,courier_id.asc,route_id.asc",
        )
        if rows:
            for row in rows:
                row["source_table"] = table_name
            return table_name, rows

    return "", []


def latest_work_date(rows: list[dict[str, Any]]) -> str:
    dates = sorted(
        {
            str(row.get("work_date") or "")[:10]
            for row in rows
            if str(row.get("work_date") or "")[:10]
        }
    )
    return dates[-1] if dates else ""


def parse_raw_attendance_rows(table_name: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []

    for raw in rows:
        work_date = str(raw.get("work_date") or "")[:10]
        payload = raw.get("response_json") or {}
        if not isinstance(payload, dict):
            continue

        for courier in payload.get("couriers") or []:
            if not isinstance(courier, dict):
                continue

            courier_id = to_int(courier.get("courierId"))
            courier_name = str(courier.get("courierName") or "").strip()
            warehouse_name = str(courier.get("warehouseName") or "").strip()

            for route in courier.get("routes") or []:
                if not isinstance(route, dict):
                    continue

                route_id = to_int(route.get("routeId") or route.get("id"))
                if route_id is None:
                    continue

                planned_departure = route.get("plannedDeparture")
                real_departure = route.get("realDeparture")
                planned_return = route.get("plannedReturn")
                real_return = route.get("realReturn")

                output.append(
                    {
                        "work_date": work_date,
                        "courier_id": courier_id,
                        "courier_name": courier_name,
                        "warehouse_name": warehouse_name,
                        "route_id": route_id,
                        "courier_registered_at": route.get("courierRegisteredAt"),
                        "assigned_at": route.get("assignedAt"),
                        "planned_departure": planned_departure,
                        "real_departure": real_departure,
                        "planned_return": planned_return,
                        "real_return": real_return,
                        "planned_route_minutes": minutes_between(planned_return, planned_departure),
                        "real_route_minutes": minutes_between(real_return, real_departure),
                        "departure_diff_minutes": minutes_between(real_departure, planned_departure),
                        "return_diff_minutes": minutes_between(real_return, planned_return),
                        "departure_status": "",
                        "return_status": "",
                        "source_table": table_name,
                    }
                )

    return output


def read_raw_rows(
    *,
    start_date: str,
    end_date: str,
    page_size: int,
) -> tuple[str, list[dict[str, Any]]]:
    for table_name in RAW_TABLE_CANDIDATES:
        rows = fetch_table_rows(
            table_name,
            select="work_date,response_json",
            start_date=start_date,
            end_date=end_date,
            page_size=page_size,
            order="work_date.asc",
        )
        if rows:
            parsed = parse_raw_attendance_rows(table_name, rows)
            if parsed:
                return table_name, parsed

    return "", []


def clean_cell_value(column: str, value: Any) -> Any:
    if column in DATE_COLUMNS:
        return parse_date_value(value)
    if column in DATETIME_COLUMNS:
        return parse_datetime_value(value)
    if column in NUMBER_COLUMNS:
        return to_int(value)
    return value


def style_sheet(ws, column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="163B2B")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, date) and not isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"

    for column_index in range(1, column_count + 1):
        column_letter = get_column_letter(column_index)
        max_length = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)), 42))
        ws.column_dimensions[column_letter].width = max_length + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_routes_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Route ID lista")
    ws.append([HEADERS_HU.get(column, column) for column in EXPORT_COLUMNS])
    for row in rows:
        ws.append([clean_cell_value(column, row.get(column)) for column in EXPORT_COLUMNS])
    style_sheet(ws, len(EXPORT_COLUMNS))


def add_summary_sheet(wb: Workbook, rows: list[dict[str, Any]], period_label: str) -> None:
    summary: dict[int, dict[str, Any]] = {}

    for row in rows:
        courier_id = to_int(row.get("courier_id")) or 0
        route_id = to_int(row.get("route_id"))
        item = summary.setdefault(
            courier_id,
            {
                "courier_id": courier_id,
                "courier_name": row.get("courier_name") or "",
                "warehouse_names": set(),
                "route_ids": set(),
                "work_dates": set(),
            },
        )
        if row.get("warehouse_name"):
            item["warehouse_names"].add(str(row.get("warehouse_name")))
        if route_id is not None:
            item["route_ids"].add(route_id)
        if row.get("work_date"):
            item["work_dates"].add(str(row.get("work_date"))[:10])

    ws = wb.create_sheet("Osszesito")
    ws.append(["Honap", "Futar ID", "Futar neve", "Raktarak", "Route db", "Nap db"])
    for item in sorted(summary.values(), key=lambda row: str(row["courier_name"])):
        ws.append(
            [
                period_label,
                item["courier_id"] or "",
                item["courier_name"],
                ", ".join(sorted(item["warehouse_names"])),
                len(item["route_ids"]),
                len(item["work_dates"]),
            ]
        )
    style_sheet(ws, 6)


def add_meta_sheet(
    wb: Workbook,
    *,
    start_date: str,
    end_date: str,
    source_table: str,
    row_count: int,
) -> None:
    ws = wb.create_sheet("Meta")
    ws.append(["Mezo", "Ertek"])
    ws.append(["Idoszak", f"{start_date} - {end_date}"])
    ws.append(["Forras", f"public.{source_table}"])
    ws.append(["Exportalt route sor", row_count])
    ws.append(["Generalva", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    style_sheet(ws, 2)


def write_workbook(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    period_label: str,
    start_date: str,
    end_date: str,
    source_table: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_routes_sheet(wb, rows)
    add_summary_sheet(wb, rows, period_label)
    add_meta_sheet(
        wb,
        start_date=start_date,
        end_date=end_date,
        source_table=source_table,
        row_count=len(rows),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    period_label, start_date, end_date = resolve_dates(args)
    output_path = Path(
        args.output
        or f"exports/attendance_route_ids_{period_label}.xlsx"
    )

    source_table = ""
    rows: list[dict[str, Any]] = []

    if args.source in ("auto", "routes"):
        source_table, rows = read_route_table_rows(
            start_date=start_date,
            end_date=end_date,
            page_size=args.page_size,
        )

    route_latest = latest_work_date(rows)
    route_table_is_full = bool(rows) and route_latest >= end_date

    if args.source == "raw" or not rows or (
        args.source == "auto" and not route_table_is_full
    ):
        if rows and args.source == "auto":
            print(
                "Figyelmeztetes: a bontott attendance route tabla hianyos "
                f"(legutolso datum: {route_latest}, vart: {end_date}), raw JSON-bol exportalok."
            )
        source_table, rows = read_raw_rows(
            start_date=start_date,
            end_date=end_date,
            page_size=args.page_size,
        )

    if not rows:
        raise RuntimeError("Nem talaltam route ID sort az adott idoszakra.")

    rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("work_date") or ""),
            str(row.get("courier_name") or ""),
            to_int(row.get("route_id")) or 0,
        ),
    )
    write_workbook(
        rows=rows,
        output_path=output_path,
        period_label=period_label,
        start_date=start_date,
        end_date=end_date,
        source_table=source_table,
    )
    print(f"Kesz: {output_path} | route sorok: {len(rows)} | forras: {source_table}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
