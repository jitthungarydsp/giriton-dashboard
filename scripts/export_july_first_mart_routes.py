#!/usr/bin/env python3
"""
Juliusi elso napi mart route export Excelbe.

Pelda:
    python scripts/export_july_first_mart_routes.py
    python scripts/export_july_first_mart_routes.py --output exports/julius_elso_turak.xlsx
    python scripts/export_july_first_mart_routes.py --all-couriers
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dsp_incremental_common import (
    get_supabase_config,
    raise_for_supabase_error,
    supabase_headers,
)


TABLE_NAME = "mart_dsp_route_stories"
PAGE_SIZE = 500

DEFAULT_START_DATE = "2026-07-01"
DEFAULT_END_DATE = "2026-07-31"

COURIER_NAMES = [
    "Hohmann Andras Armin",
    "Zoranyi Tamas Bence",
    "Gurzo Balazs",
    "Nadasi Martin Nicolas",
    "Szolnoki Tibor",
    "Petrik Robert",
    "Herkovics Andrea",
    "Kiraly Daniel David",
    "Baranyi Szabolcs",
    "Farkas Zsolt",
    "Bari Anett",
    "Molnar Akos",
    "Bosko 8031 Gyorgy",
    "Buban Szimonetta",
    "Varga Romano",
    "Rezsnyak Bence Laszlo",
    "Mocsary Attila",
    "Gemes Zoltan",
    "Bognar Andras",
    "Adam Erik",
    "Stumpf 8055 Jozsef",
    "Mate 8032 Daniel",
    "Nagy 7725 Istvan",
    "Takacs 8223 Attila",
    "Szikszai 8266 Ferenc",
]

SOURCE_COLUMNS = [
    "courier_name",
    "courier_id",
    "work_date",
    "warehouse_name",
    "route_id",
    "address_count",
    "shift_name",
    "shift_start",
    "shift_end",
    "available_at",
    "available_for_shift_since",
    "queue_started_at",
    "assigned_at",
    "planned_departure",
    "real_departure",
    "planned_return",
    "real_return",
    "planned_route_minutes",
    "real_route_minutes",
    "assigned_to_return_minutes",
    "total_route_minutes",
    "queue_entry_delta_minutes",
    "queue_wait_minutes",
    "time_window_late_count",
    "next_shift_delay_minutes",
    "assignment_mode",
    "story_text",
]

EXPORT_COLUMNS = [
    "courier_name",
    "courier_id",
    "work_date",
    "warehouse_name",
    "route_id",
    "address_count",
    "shift_name",
    "shift_start",
    "shift_end",
    "available_at",
    "available_for_shift_since",
    "queue_started_at",
    "assigned_at",
    "real_departure",
    "planned_return",
    "real_return",
    "planned_route_minutes",
    "real_route_minutes",
    "assigned_to_return_minutes",
    "total_route_minutes",
    "queue_entry_delta_minutes",
    "queue_wait_minutes",
    "time_window_late_count",
    "next_shift_delay_minutes",
    "assignment_mode",
    "story_text",
]

HEADERS_HU = {
    "courier_name": "Futar",
    "courier_id": "Futar ID",
    "work_date": "Datum",
    "warehouse_name": "Raktar",
    "route_id": "Route ID",
    "address_count": "Cim darabszam",
    "shift_name": "Muszak",
    "shift_start": "Muszak kezdete",
    "shift_end": "Muszak vege",
    "available_at": "Elerheto volt",
    "available_for_shift_since": "Muszakhoz elerheto",
    "queue_started_at": "Sorba allt",
    "assigned_at": "Turat kapott",
    "real_departure": "Valos indulas",
    "planned_return": "Tervezett visszaerkezes",
    "real_return": "Valos visszaerkezes",
    "planned_route_minutes": "Tervezett turahossz (perc)",
    "real_route_minutes": "Turahossz (perc)",
    "assigned_to_return_minutes": "Bepakolassal turahossz (perc)",
    "total_route_minutes": "Teljes turaido (perc)",
    "queue_entry_delta_minutes": "Sorbaallas elteres (perc)",
    "queue_wait_minutes": "Varakozas turara (perc)",
    "time_window_late_count": "Idoablakhoz kepest keses (db)",
    "next_shift_delay_minutes": "Kovetkezo muszak keses (perc)",
    "assignment_mode": "Kiosztas modja",
    "story_text": "Route story",
}

DATE_COLUMNS = {"work_date"}
DATETIME_COLUMNS = {
    "shift_start",
    "shift_end",
    "available_at",
    "available_for_shift_since",
    "queue_started_at",
    "assigned_at",
    "real_departure",
    "planned_return",
    "real_return",
}
NUMBER_COLUMNS = {
    "courier_id",
    "route_id",
    "address_count",
    "planned_route_minutes",
    "real_route_minutes",
    "assigned_to_return_minutes",
    "total_route_minutes",
    "queue_entry_delta_minutes",
    "queue_wait_minutes",
    "time_window_late_count",
    "next_shift_delay_minutes",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Juliusi elso napi mart route export Excelbe."
    )
    parser.add_argument("--start-date", default=DEFAULT_START_DATE)
    parser.add_argument("--end-date", default=DEFAULT_END_DATE)
    parser.add_argument("--output", help="Kimeneti Excel fajl")
    parser.add_argument("--page-size", type=int, default=PAGE_SIZE)
    parser.add_argument(
        "--all-couriers",
        action="store_true",
        help="Ne csak a megadott 25 futart exportalja.",
    )
    return parser.parse_args()


def validate_iso_date(value: str, field_name: str) -> None:
    try:
        date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            f"Hibas {field_name}: {value!r}. Elvart formatum: EEEE-HH-NN."
        ) from exc


def normalize_name(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def missing_column_from_response(response: requests.Response) -> str | None:
    if response.status_code != 400:
        return None

    try:
        payload = response.json()
    except ValueError:
        return None

    if not isinstance(payload, dict) or payload.get("code") != "42703":
        return None

    message = str(payload.get("message") or "")
    match = re.search(r"column\s+\S+\.([A-Za-z_][A-Za-z0-9_]*)\s+does not exist", message)
    if match:
        return match.group(1)
    return None


def fetch_rows(
    *,
    start_date: str,
    end_date: str,
    page_size: int,
) -> list[dict[str, Any]]:
    supabase_url, _service_role_key = get_supabase_config()
    endpoint = f"{supabase_url}/rest/v1/{TABLE_NAME}"
    rows: list[dict[str, Any]] = []
    offset = 0
    select_columns = list(SOURCE_COLUMNS)

    while True:
        params: list[tuple[str, str]] = [
            ("select", ",".join(select_columns)),
            ("work_date", f"gte.{start_date}"),
            ("work_date", f"lte.{end_date}"),
            ("order", "work_date.asc,courier_id.asc,shift_start.asc,assigned_at.asc,route_id.asc"),
            ("limit", str(page_size)),
            ("offset", str(offset)),
        ]
        response = requests.get(
            endpoint,
            headers=supabase_headers({"Accept": "application/json"}),
            params=params,
            timeout=120,
        )

        missing_column = missing_column_from_response(response)
        if missing_column and missing_column in select_columns:
            select_columns.remove(missing_column)
            rows = []
            offset = 0
            print(
                f"Figyelmeztetes: public.{TABLE_NAME}.{missing_column} oszlop nincs, kihagyom."
            )
            continue

        raise_for_supabase_error(response, TABLE_NAME)
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError(f"Varatlan Supabase valasz: {page!r}")

        rows.extend(page)
        print(f"Letoltve: {len(page):>4} sor | osszesen: {len(rows):>6} sor")

        if len(page) < page_size:
            break
        offset += page_size

    return rows


def parse_sort_datetime(value: Any) -> str:
    if value in (None, ""):
        return "9999-99-99T99:99:99"
    return str(value)


def parse_sort_int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def first_route_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        str(row.get("work_date") or ""),
        parse_sort_int(row.get("courier_id")),
        parse_sort_datetime(row.get("shift_start")),
        parse_sort_datetime(row.get("assigned_at")),
        parse_sort_int(row.get("route_id")),
    )


def build_first_route_rows(
    rows: list[dict[str, Any]],
    *,
    all_couriers: bool,
) -> list[dict[str, Any]]:
    allowed_names = {normalize_name(name) for name in COURIER_NAMES}
    first_by_day: dict[tuple[int, str], dict[str, Any]] = {}

    for row in sorted(rows, key=first_route_sort_key):
        if not all_couriers and normalize_name(row.get("courier_name")) not in allowed_names:
            continue

        courier_id = parse_sort_int(row.get("courier_id"))
        work_date = str(row.get("work_date") or "")[:10]
        if not courier_id or not work_date:
            continue

        key = (courier_id, work_date)
        if key not in first_by_day:
            first_by_day[key] = row

    return sorted(first_by_day.values(), key=lambda row: (str(row.get("courier_name") or ""), str(row.get("work_date") or "")))


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def parse_datetime_value(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
        except ValueError:
            return None
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt


def clean_cell_value(column: str, value: Any) -> Any:
    if column in DATE_COLUMNS:
        return parse_date_value(value)
    if column in DATETIME_COLUMNS:
        return parse_datetime_value(value)
    if column in NUMBER_COLUMNS and value not in (None, ""):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return value
    return value


def add_rows_sheet(wb: Workbook, rows: list[dict[str, Any]], title: str) -> None:
    ws = wb.create_sheet(title)
    ws.append([HEADERS_HU.get(column, column) for column in EXPORT_COLUMNS])

    for row in rows:
        ws.append([clean_cell_value(column, row.get(column)) for column in EXPORT_COLUMNS])

    style_sheet(ws, len(EXPORT_COLUMNS))
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[int, dict[str, Any]] = {}

    for row in rows:
        courier_id = parse_sort_int(row.get("courier_id"))
        if not courier_id:
            continue

        item = summary.setdefault(
            courier_id,
            {
                "courier_name": row.get("courier_name"),
                "courier_id": courier_id,
                "day_count": 0,
                "late_route_days": 0,
                "late_route_count": 0,
                "address_count": 0,
            },
        )
        item["day_count"] += 1
        late_count = parse_sort_int(row.get("time_window_late_count"))
        item["late_route_count"] += late_count
        if late_count > 0:
            item["late_route_days"] += 1
        item["address_count"] += parse_sort_int(row.get("address_count"))

    return sorted(summary.values(), key=lambda row: str(row.get("courier_name") or ""))


def add_summary_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    columns = [
        ("courier_name", "Futar"),
        ("courier_id", "Futar ID"),
        ("day_count", "Elso tura napok"),
        ("address_count", "Cimek osszesen"),
        ("late_route_days", "Idoablakos keses napok"),
        ("late_route_count", "Idoablakos keses cimek"),
    ]
    ws = wb.create_sheet("Osszesito")
    ws.append([label for _key, label in columns])
    for row in build_summary_rows(rows):
        ws.append([row.get(key) for key, _label in columns])
    style_sheet(ws, len(columns))
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def add_meta_sheet(
    wb: Workbook,
    *,
    start_date: str,
    end_date: str,
    row_count: int,
    all_couriers: bool,
) -> None:
    ws = wb.create_sheet("Meta")
    ws.append(["Mezo", "Ertek"])
    ws.append(["Idoszak", f"{start_date} - {end_date}"])
    ws.append(["Forras", "public.mart_dsp_route_stories"])
    ws.append(["Logika", "courier_id + work_date szerint az elso shift_start / assigned_at / route_id sor"])
    ws.append(["Exportalt sor", row_count])
    ws.append(["Futar szures", "Minden futar" if all_couriers else "Megadott 25 futar"])
    ws.append(["Generalva", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    style_sheet(ws, 2)


def style_sheet(ws, column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="163B2B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, date) and not isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"

    for column_index in range(1, column_count + 1):
        column_letter = get_column_letter(column_index)
        max_length = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)), 45))
        ws.column_dimensions[column_letter].width = max_length + 2


def write_workbook(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    start_date: str,
    end_date: str,
    all_couriers: bool,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_rows_sheet(wb, rows, "Elso turak")
    add_summary_sheet(wb, rows)
    add_meta_sheet(
        wb,
        start_date=start_date,
        end_date=end_date,
        row_count=len(rows),
        all_couriers=all_couriers,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    validate_iso_date(args.start_date, "start-date")
    validate_iso_date(args.end_date, "end-date")

    output_path = Path(
        args.output
        or f"exports/elso_turak_mart_{args.start_date}_{args.end_date}.xlsx"
    )

    raw_rows = fetch_rows(
        start_date=args.start_date,
        end_date=args.end_date,
        page_size=args.page_size,
    )
    output_rows = build_first_route_rows(raw_rows, all_couriers=args.all_couriers)
    write_workbook(
        rows=output_rows,
        output_path=output_path,
        start_date=args.start_date,
        end_date=args.end_date,
        all_couriers=args.all_couriers,
    )

    print(f"Kesz: {output_path} | sorok: {len(output_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
