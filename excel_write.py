from openpyxl import Workbook, load_workbook
from os.path import exists
from dsp_common_kw import hu_time
def write_shift(file_name, datum, kezdes, vege, raktar, foglaltsag, nev):

    if not exists(file_name):
        wb = Workbook()
        ws = wb.active

        ws.append([
            "Dátum",
            "Kezdés",
            "Vége",
            "Raktár",
            "Foglaltság",
            "Név"
        ])

        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    ws.append([
        datum,
        kezdes,
        vege,
        raktar,
        foglaltsag,
        nev
    ])

    wb.save(file_name)